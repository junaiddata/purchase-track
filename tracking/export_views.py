"""
Export views for PDF and Excel downloads.
"""
import re

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum

from .models import Release, QuotationItem, ItemMaster
from .decorators import admin_required, sales_required
from .utils import fetch_local_open_qty_map


def parse_consolidated_search_tokens(raw):
    """
    Same rules as Purchase Report / sales_firm_track UI: newline, comma, or semicolon
    separates codes; space-separated tokens only if every part looks code-like ([\\w.-]+).
    """
    if raw is None:
        return []
    trimmed = str(raw).strip()
    if not trimmed:
        return []
    parts = [p.strip() for p in re.split(r"[\n,;]+", trimmed) if p.strip()]
    if len(parts) > 1:
        return [p.lower() for p in parts]
    single = parts[0]
    by_space = [p for p in single.split() if p]
    if len(by_space) > 1 and all(re.match(r"^[\w.-]+$", p) for p in by_space):
        return [p.lower() for p in by_space]
    return [single.lower()]


def item_matches_consolidated_tokens(item_code, item_description, tokens):
    if not tokens:
        return True
    code_l = (item_code or "").lower()
    desc_l = (item_description or "").lower()
    if len(tokens) == 1:
        t = tokens[0]
        return t in code_l or t in desc_l
    return any(t in code_l or t in desc_l for t in tokens)


def _key_matches_single_token(key, t):
    code_l = key[0].lower()
    desc_l = (key[1] or "").lower()
    return t in code_l or t in desc_l


def ordered_consolidated_keys(keys, tokens):
    """Row order: groups matching each token in search order, then remaining keys sorted by item code."""
    if not keys:
        return keys
    keys_sorted = sorted(keys)
    if not tokens or len(tokens) <= 1:
        return keys_sorted
    seen = set()
    ordered = []
    for t in tokens:
        if not t:
            continue
        for k in keys_sorted:
            if k in seen or not _key_matches_single_token(k, t):
                continue
            ordered.append(k)
            seen.add(k)
    for k in keys_sorted:
        if k not in seen:
            ordered.append(k)
    return ordered


@login_required
@sales_required
def export_sales_track_pdf(request):
    """
    Export sales tracking view to PDF.
    By default no rates/amounts. Admins can toggle show_rates=1.
    
    Self-contained module — all helpers are local, no cross-app PDF dependencies.
    """
    from datetime import datetime
    from io import BytesIO
    import re
    import os

    import requests
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.lib.units import inch, mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether,
    )

    # ─────────────────────────────────────────────────────────────
    # DESIGN TOKENS — single source of truth for this report
    # ─────────────────────────────────────────────────────────────

    # Brand palette
    CLR_NAVY        = HexColor('#0F172A')    # Primary dark — headers, titles
    CLR_NAVY_MED    = HexColor('#1E293B')    # Section bar background
    CLR_SLATE       = HexColor('#334155')    # Body text
    CLR_SLATE_LIGHT = HexColor('#64748B')    # Muted / meta text
    CLR_ACCENT      = HexColor('#0EA5E9')    # Sky blue — accent lines, markers
    CLR_ACCENT_DARK = HexColor('#0284C7')    # Darker accent for emphasis
    CLR_BORDER      = HexColor('#CBD5E1')    # Table grid — subtle
    CLR_BORDER_HEAVY= HexColor('#94A3B8')    # Outer box
    CLR_ROW_ALT     = HexColor('#F8FAFC')    # Zebra stripe
    CLR_ROW_WHITE   = colors.white
    CLR_HEADER_BG   = HexColor('#0F172A')    # Table header background
    CLR_HEADER_FG   = colors.white
    CLR_TOTAL_BG    = HexColor('#EFF6FF')    # Summary row tint
    CLR_SECTION_BG  = HexColor('#F1F5F9')    # Section header bar bg
    CLR_RECEIVED    = HexColor('#059669')    # Green for "Delivered"
    CLR_PENDING     = HexColor('#D97706')    # Amber for pending
    CLR_TRANSIT     = HexColor('#2563EB')    # Blue for in-transit
    CLR_WHITE       = colors.white
    CLR_RED_SOFT    = HexColor('#FEF2F2')

    # Typography
    FONT_TITLE      = 16
    FONT_SUBTITLE   = 9
    FONT_SECTION    = 10.5
    FONT_TH         = 7.5
    FONT_TD         = 7
    FONT_TD_SM      = 6.5
    FONT_META       = 8
    FONT_FOOTER     = 6.5
    FONT_BADGE      = 6.5

    # Spacing
    SP_SECTION      = 14
    SP_AFTER_TABLE  = 12
    SP_INNER        = 5

    # Cell padding
    PAD_H           = 5
    PAD_V_HDR       = 7
    PAD_V_DATA      = 4.5

    # ─────────────────────────────────────────────────────────────
    # LOCAL STYLE FACTORY
    # ─────────────────────────────────────────────────────────────

    base_styles = getSampleStyleSheet()
    _base = base_styles['Normal']

    def _ps(name, **kwargs):
        """Shorthand ParagraphStyle constructor."""
        return ParagraphStyle(name, parent=_base, **kwargs)

    S_TITLE = _ps('RptTitle',
        fontName='Helvetica-Bold', fontSize=FONT_TITLE,
        textColor=CLR_NAVY, leading=FONT_TITLE + 4, alignment=TA_LEFT,
    )
    S_TITLE_R = _ps('RptTitleR',
        fontName='Helvetica-Bold', fontSize=FONT_TITLE,
        textColor=CLR_NAVY, leading=FONT_TITLE + 4, alignment=TA_RIGHT,
    )
    S_SUBTITLE = _ps('RptSub',
        fontName='Helvetica', fontSize=FONT_SUBTITLE,
        textColor=CLR_SLATE_LIGHT, leading=FONT_SUBTITLE + 3,
    )
    S_SUBTITLE_R = _ps('RptSubR',
        fontName='Helvetica', fontSize=FONT_SUBTITLE,
        textColor=CLR_SLATE_LIGHT, leading=FONT_SUBTITLE + 3, alignment=TA_RIGHT,
    )
    S_SECTION = _ps('RptSection',
        fontName='Helvetica-Bold', fontSize=FONT_SECTION,
        textColor=CLR_NAVY, leading=FONT_SECTION + 3,
    )
    S_TH = _ps('RptTH',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 3,
    )
    S_TH_R = _ps('RptTHR',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 3, alignment=TA_RIGHT,
    )
    S_TH_C = _ps('RptTHC',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 3, alignment=TA_CENTER,
    )
    S_TD = _ps('RptTD',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 3, wordWrap='CJK',
    )
    S_TD_C = _ps('RptTDC',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 3, alignment=TA_CENTER,
    )
    S_TD_R = _ps('RptTDR',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 3, alignment=TA_RIGHT,
    )
    S_TD_BOLD = _ps('RptTDB',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 3,
    )
    S_TD_BOLD_R = _ps('RptTDBR',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 3, alignment=TA_RIGHT,
    )
    S_META = _ps('RptMeta',
        fontName='Helvetica', fontSize=FONT_META,
        textColor=CLR_SLATE_LIGHT, leading=FONT_META + 3,
    )
    S_BADGE_DELIVERED = _ps('BadgeDel',
        fontName='Helvetica-Bold', fontSize=FONT_BADGE,
        textColor=CLR_RECEIVED, leading=FONT_BADGE + 2, alignment=TA_CENTER,
    )
    S_BADGE_TRANSIT = _ps('BadgeTr',
        fontName='Helvetica-Bold', fontSize=FONT_BADGE,
        textColor=CLR_TRANSIT, leading=FONT_BADGE + 2, alignment=TA_CENTER,
    )
    S_BADGE_PENDING = _ps('BadgePn',
        fontName='Helvetica-Bold', fontSize=FONT_BADGE,
        textColor=CLR_PENDING, leading=FONT_BADGE + 2, alignment=TA_CENTER,
    )
    S_EMPTY = _ps('RptEmpty',
        fontName='Helvetica-Oblique', fontSize=FONT_TD,
        textColor=CLR_SLATE_LIGHT, leading=FONT_TD + 3, alignment=TA_CENTER,
    )

    # ─────────────────────────────────────────────────────────────
    # LOCAL HELPER FUNCTIONS
    # ─────────────────────────────────────────────────────────────

    def _esc(text):
        """Escape text for ReportLab Paragraph XML."""
        s = str(text) if text is not None else ''
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def _p(text, style=None):
        """Create a Paragraph with escaped text."""
        return Paragraph(_esc(text), style or S_TD)

    def _p_num(value, style=None):
        """Format a number and wrap in a right-aligned Paragraph."""
        try:
            v = float(value) if value else 0
            formatted = f"{v:,.0f}" if v == int(v) else f"{v:,.2f}"
        except (TypeError, ValueError):
            formatted = str(value or '0')
        return Paragraph(formatted, style or S_TD_R)

    def _p_currency(amount, currency_display, style=None):
        """Format currency value: 'USD 1,234.56'"""
        try:
            v = float(amount) if amount else 0
            formatted = f"{currency_display} {v:,.2f}"
        except (TypeError, ValueError):
            formatted = f"{currency_display} 0.00"
        return Paragraph(_esc(formatted), style or S_TD_R)

    def _p_date(date_val, fallback='TBD', style=None):
        """Format a date or show fallback."""
        if date_val:
            try:
                return Paragraph(date_val.strftime('%d %b %Y'), style or S_TD_C)
            except Exception:
                pass
        return Paragraph(fallback, style or S_TD_C)

    def _load_logo():
        """Load logo: local first, then URL fallback. Returns bytes or None."""
        # Local candidates
        local_paths = [
            os.path.join(settings.BASE_DIR, 'media', 'footer-logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'footer-logo1.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'footer-logo.png'),
        ]
        for path in local_paths:
            if os.path.exists(path):
                try:
                    with open(path, 'rb') as f:
                        return f.read()
                except Exception:
                    continue
        # URL fallback
        try:
            url = 'https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp'
            r = requests.get(url, timeout=5)
            if r.status_code == 200 and r.content:
                return r.content
        except Exception:
            pass
        return None

    def _build_section_bar(number, title, page_width):
        """
        Build a section header: accent left-strip + number badge + title.
        Returns a Table flowable.
        """
        accent_w = 4
        badge_w = 0.35 * inch
        content_w = page_width - accent_w - badge_w - 1.0 * inch

        # Number badge
        badge_style = _ps(f'Badge{number}',
            fontName='Helvetica-Bold', fontSize=9,
            textColor=CLR_WHITE, alignment=TA_CENTER,
        )
        # Title
        title_para = Paragraph(title, S_SECTION)
        badge_para = Paragraph(str(number), badge_style)

        row = ['', badge_para, title_para]
        tbl = Table([row], colWidths=[accent_w, badge_w, content_w])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), CLR_ACCENT),
            ('BACKGROUND', (1, 0), (1, 0), CLR_ACCENT),
            ('BACKGROUND', (2, 0), (2, 0), CLR_SECTION_BG),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (2, 0), (2, 0), 8),
            # Round effect: subtle bottom border
            ('LINEBELOW', (0, 0), (-1, -1), 0.5, CLR_BORDER),
        ]))
        return tbl

    def _build_data_table(headers, data_rows, col_widths, num_right_align_from=None):
        """
        Build a professional data table with:
        - Dark header row
        - Zebra striping
        - Subtle grid
        - Proper alignment
        
        headers: list of (text, style) tuples
        data_rows: list of lists (already Paragraph objects)
        col_widths: list of column widths
        num_right_align_from: column index from which to right-align (for numeric cols)
        """
        # Build header row
        header_row = [Paragraph(_esc(h[0]), h[1]) for h in headers]
        all_rows = [header_row] + data_rows

        tbl = Table(all_rows, colWidths=col_widths, repeatRows=1)

        num_rows = len(all_rows)
        cmds = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), CLR_HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), CLR_HEADER_FG),

            # Outer border
            ('BOX', (0, 0), (-1, -1), 0.75, CLR_BORDER_HEAVY),
            ('LINEBELOW', (0, 0), (-1, 0), 1, CLR_BORDER_HEAVY),

            # Cell padding — header
            ('TOPPADDING', (0, 0), (-1, 0), PAD_V_HDR),
            ('BOTTOMPADDING', (0, 0), (-1, 0), PAD_V_HDR),

            # Cell padding — data
            ('TOPPADDING', (0, 1), (-1, -1), PAD_V_DATA),
            ('BOTTOMPADDING', (0, 1), (-1, -1), PAD_V_DATA),
            ('LEFTPADDING', (0, 0), (-1, -1), PAD_H),
            ('RIGHTPADDING', (0, 0), (-1, -1), PAD_H),

            # Alignment
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Zebra striping + subtle row lines
        for i in range(1, num_rows):
            bg = CLR_ROW_ALT if i % 2 == 0 else CLR_ROW_WHITE
            cmds.append(('BACKGROUND', (0, i), (-1, i), bg))
            if i < num_rows - 1:
                cmds.append(('LINEBELOW', (0, i), (-1, i), 0.25, CLR_BORDER))

        # Last data row gets a slightly heavier bottom line
        if num_rows > 1:
            cmds.append(('LINEBELOW', (0, -1), (-1, -1), 0.5, CLR_BORDER_HEAVY))

        tbl.setStyle(TableStyle(cmds))
        return tbl

    def _empty_row(col_count, message="No data available"):
        """Build a single empty-state row."""
        cells = [Paragraph('', S_TD)] * col_count
        cells[0] = Paragraph(message, S_EMPTY)
        # Span across all columns
        return cells

    def _build_count_badge(count, label, color):
        """Build an inline count indicator: '12 items'"""
        return Paragraph(
            f'<font color="{color}"><b>{count}</b></font> '
            f'<font color="#94A3B8" size="7">{label}</font>',
            _ps('CountBadge', fontSize=8, leading=10),
        )

    # ─────────────────────────────────────────────────────────────
    # BUSINESS LOGIC (unchanged)
    # ─────────────────────────────────────────────────────────────

    firm_name = request.GET.get('firm')
    show_rates = request.GET.get('show_rates', '').lower() in ('1', 'true', 'yes')
    show_received = request.GET.get('show_received', '1').lower() in ('1', 'true', 'yes')
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'

    if not is_admin:
        show_rates = False

    if not firm_name:
        return HttpResponse("Firm name required", status=400)

    logo_bytes = _load_logo()

    in_transit_releases = Release.objects.filter(
        quotation_item__item__item_firm=firm_name,
        is_received=False
    ).select_related(
        'quotation_item__item',
        'quotation_item__quotation',
        'quotation_item__quotation__manufacturer'
    ).order_by('container_info', 'expected_arrival_date')

    received_queryset = Release.objects.filter(
        quotation_item__item__item_firm=firm_name,
        is_received=True
    ).select_related(
        'quotation_item__item',
        'quotation_item__quotation',
        'quotation_item__quotation__manufacturer'
    ).prefetch_related('quotation_item__shipments').order_by('-release_date')

    all_firm_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status='CONFIRMED'
    ).select_related('item', 'quotation', 'quotation__manufacturer').order_by('expected_delivery_date')

    pending_items = [item for item in all_firm_items if item.balance_to_release > 0]

    for release in in_transit_releases:
        release.transit_amount = release.quantity_released * release.quotation_item.rate
        release.current_stock = release.quotation_item.item.item_stock
        release.sold_stock = release.quotation_item.item.total_qty if is_admin else None

    received_list = list(received_queryset)
    if show_received:
        for release in received_list:
            release.received_amount = release.quantity_released * release.quotation_item.rate
            shipment = release.quotation_item.shipments.filter(
                quantity_received=release.quantity_released
            ).order_by('-received_date').first()
            release.actual_received_date = shipment.received_date if shipment else release.release_date
            release.current_stock = release.quotation_item.item.item_stock
            release.sold_stock = release.quotation_item.item.total_qty if is_admin else None

    for item in pending_items:
        item.balance_amount = item.balance_to_release * item.rate
        item.current_stock = item.item.item_stock
        item.sold_stock = item.item.total_qty if is_admin else None

    # ─────────────────────────────────────────────────────────────
    # PDF DOCUMENT SETUP
    # ─────────────────────────────────────────────────────────────

    buffer = BytesIO()
    pw, ph = landscape(A4)
    margin_h = 0.5 * inch
    margin_top = 0.7 * inch
    margin_bot = 0.55 * inch
    usable_w = pw - 2 * margin_h

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=margin_h, rightMargin=margin_h,
        topMargin=margin_top, bottomMargin=margin_bot,
    )

    elements = []

    # ─────────────────────────────────────────────────────────────
    # 1. DOCUMENT HEADER
    # ─────────────────────────────────────────────────────────────

    # Right-side info block
    title_block = Paragraph(f"<b>Import Purchase Report</b>", S_TITLE_R)
    firm_block = Paragraph(_esc(firm_name), _ps('FirmName',
        fontName='Helvetica', fontSize=12,
        textColor=CLR_ACCENT_DARK, leading=15, alignment=TA_RIGHT,
    ))
    date_block = Paragraph(
        f"Generated: {datetime.now().strftime('%d %B %Y  •  %H:%M')}",
        S_SUBTITLE_R,
    )

    right_stack = Table(
        [[title_block], [firm_block], [date_block]],
        colWidths=[usable_w - 2.0 * inch],
    )
    right_stack.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    if logo_bytes:
        header_row = [['', right_stack]]
        header_widths = [2.0 * inch, usable_w - 2.0 * inch]
    else:
        # Text-only fallback
        brand = Paragraph('<b>JUNAID</b>', _ps('Brand',
            fontName='Helvetica-Bold', fontSize=18, textColor=CLR_NAVY,
        ))
        header_row = [[brand, right_stack]]
        header_widths = [2.0 * inch, usable_w - 2.0 * inch]

    header_table = Table(header_row, colWidths=header_widths)
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 4))

    # Accent divider
    divider = Table([['']], colWidths=[usable_w])
    divider.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, CLR_ACCENT),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(divider)
    elements.append(Spacer(1, 6))

    # Summary badges row
    badges = [
        _build_count_badge(in_transit_releases.count(), 'In Transit', '#2563EB'),
    ]
    if show_received:
        badges.append(_build_count_badge(min(len(received_list), 100), 'Received', '#059669'))
    badges.append(_build_count_badge(len(pending_items), 'Pending', '#D97706'))

    badge_tbl = Table([badges], colWidths=[usable_w / len(badges)] * len(badges))
    badge_tbl.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, -1), CLR_SECTION_BG),
        ('BOX', (0, 0), (-1, -1), 0.5, CLR_BORDER),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, CLR_BORDER),
    ]))
    elements.append(badge_tbl)
    elements.append(Spacer(1, SP_SECTION))

    # ─────────────────────────────────────────────────────────────
    # COLUMN WIDTH CALCULATOR
    # ─────────────────────────────────────────────────────────────

    def _calc_widths(base_widths, desc_index=1):
        """Adjust description column to fill remaining usable width."""
        allocated = sum(base_widths)
        remainder = usable_w - allocated
        if remainder > 0:
            base_widths[desc_index] += remainder
        return base_widths

    # ─────────────────────────────────────────────────────────────
    # 2. SECTION 1 — ON THE WAY
    # ─────────────────────────────────────────────────────────────

    elements.append(_build_section_bar(1, 'ON THE WAY — In Transit Shipments', usable_w))
    elements.append(Spacer(1, SP_INNER))

    # Dynamic headers
    headers_1 = [
        ('Item Code', S_TH),
        ('Item Description', S_TH),
        ('Container', S_TH_C),
        ('Qty', S_TH_R),
        ('Stock', S_TH_R),
    ]
    widths_1 = [55, 180, 75, 50, 45]

    if is_admin:
        headers_1.append(('Sold', S_TH_R))
        widths_1.append(42)
    if show_rates:
        headers_1.extend([('Rate', S_TH_R), ('Amount', S_TH_R)])
        widths_1.extend([60, 65])
    headers_1.append(('Expected @ Port', S_TH_C))
    widths_1.append(80)

    widths_1 = _calc_widths(widths_1, desc_index=1)

    # Data rows
    rows_1 = []
    for release in in_transit_releases:
        item = release.quotation_item.item
        currency = release.quotation_item.quotation.get_currency_display()
        row = [
            _p(item.item_code, S_TD_BOLD),
            _p(item.item_description or '—', S_TD),
            _p(release.container_info or 'TBD', S_TD_C),
            _p_num(release.quantity_released),
            _p_num(release.current_stock or 0),
        ]
        if is_admin:
            row.append(_p_num(release.sold_stock or 0))
        if show_rates:
            row.append(_p_currency(release.quotation_item.rate, currency))
            row.append(_p_currency(release.transit_amount, currency, S_TD_BOLD_R))
        row.append(_p_date(release.expected_arrival_date, 'TBD'))
        rows_1.append(row)

    if not rows_1:
        rows_1.append(_empty_row(len(headers_1), 'No items currently in transit'))

    table_1 = _build_data_table(headers_1, rows_1, widths_1)
    elements.append(table_1)
    elements.append(Spacer(1, SP_AFTER_TABLE))

    # ─────────────────────────────────────────────────────────────
    # 3. SECTION 2 — RECEIVED HISTORY (optional)
    # ─────────────────────────────────────────────────────────────

    section_num = 2
    if show_received:
        elements.append(_build_section_bar(section_num, 'RECEIVED — Shipment History', usable_w))
        elements.append(Spacer(1, SP_INNER))

        headers_2 = [
            ('Received', S_TH_C),
            ('Item Code', S_TH),
            ('Item Description', S_TH),
            ('Container', S_TH_C),
            ('Qty', S_TH_R),
            ('Stock', S_TH_R),
        ]
        widths_2 = [62, 55, 165, 70, 45, 45]

        if is_admin:
            headers_2.append(('Sold', S_TH_R))
            widths_2.append(42)
        if show_rates:
            headers_2.extend([('Rate', S_TH_R), ('Amount', S_TH_R)])
            widths_2.extend([60, 65])
        headers_2.append(('Status', S_TH_C))
        widths_2.append(55)

        widths_2 = _calc_widths(widths_2, desc_index=2)

        rows_2 = []
        for release in received_list[:100]:
            item = release.quotation_item.item
            currency = release.quotation_item.quotation.get_currency_display()
            row = [
                _p_date(release.actual_received_date, '—'),
                _p(item.item_code, S_TD_BOLD),
                _p(item.item_description or '—', S_TD),
                _p(release.container_info or '—', S_TD_C),
                _p_num(release.quantity_released),
                _p_num(release.current_stock or 0),
            ]
            if is_admin:
                row.append(_p_num(release.sold_stock or 0))
            if show_rates:
                row.append(_p_currency(release.quotation_item.rate, currency))
                row.append(_p_currency(release.received_amount, currency, S_TD_BOLD_R))
            row.append(Paragraph('✓ Delivered', S_BADGE_DELIVERED))
            rows_2.append(row)

        if not rows_2:
            rows_2.append(_empty_row(len(headers_2), 'No received shipments'))

        table_2 = _build_data_table(headers_2, rows_2, widths_2)
        elements.append(table_2)
        elements.append(Spacer(1, SP_AFTER_TABLE))
        section_num = 3

    # ─────────────────────────────────────────────────────────────
    # 4. SECTION 3 — PENDING ORDERS
    # ─────────────────────────────────────────────────────────────

    elements.append(_build_section_bar(section_num, 'PENDING — Orders at Factory', usable_w))
    elements.append(Spacer(1, SP_INNER))

    headers_3 = [
        ('Item Code', S_TH),
        ('Item Description', S_TH),
        ('Manufacturer', S_TH),
        ('Balance', S_TH_R),
        ('Stock', S_TH_R),
    ]
    widths_3 = [55, 150, 75, 52, 45]

    if is_admin:
        headers_3.append(('Sold', S_TH_R))
        widths_3.append(42)
    if show_rates:
        headers_3.extend([('Rate', S_TH_R), ('Amount', S_TH_R)])
        widths_3.extend([60, 65])
    headers_3.extend([('PO Ref', S_TH), ('Expected', S_TH_C)])
    widths_3.extend([65, 75])

    widths_3 = _calc_widths(widths_3, desc_index=1)

    rows_3 = []
    for qi in pending_items:
        currency = qi.quotation.get_currency_display()
        mfr = qi.quotation.manufacturer.name if qi.quotation.manufacturer else '—'
        row = [
            _p(qi.item.item_code, S_TD_BOLD),
            _p(qi.item.item_description or '—', S_TD),
            _p(mfr, S_TD),
            _p_num(qi.balance_to_release),
            _p_num(qi.current_stock or 0),
        ]
        if is_admin:
            row.append(_p_num(qi.sold_stock or 0))
        if show_rates:
            row.append(_p_currency(qi.rate, currency))
            row.append(_p_currency(qi.balance_amount, currency, S_TD_BOLD_R))
        row.append(_p(qi.quotation.reference_number or '—', S_TD))
        row.append(_p_date(qi.expected_delivery_date, '—'))
        rows_3.append(row)

    if not rows_3:
        rows_3.append(_empty_row(len(headers_3), 'No pending orders at factory'))

    table_3 = _build_data_table(headers_3, rows_3, widths_3)
    elements.append(table_3)

    # ─────────────────────────────────────────────────────────────
    # PAGE HEADER & FOOTER CALLBACK
    # ─────────────────────────────────────────────────────────────

    def _draw_page_chrome(canvas, doc):
        """Draw logo in header and branded footer on every page."""
        canvas.saveState()

        # ── Header logo ──
        if logo_bytes:
            try:
                img_reader = ImageReader(BytesIO(logo_bytes))
                logo_w = 1.6 * inch
                logo_h = 0.45 * inch
                logo_x = margin_h
                logo_y = ph - margin_top + 8
                canvas.drawImage(
                    img_reader, logo_x, logo_y,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto',
                )
            except Exception:
                pass

        # ── Footer ──
        footer_y = margin_bot - 20

        # Accent line
        canvas.setStrokeColor(CLR_ACCENT)
        canvas.setLineWidth(0.75)
        canvas.line(margin_h, footer_y + 12, pw - margin_h, footer_y + 12)

        # Footer text
        canvas.setFont('Helvetica', FONT_FOOTER)
        canvas.setFillColor(CLR_SLATE_LIGHT)
        footer_text = (
            f"Page {doc.page}  ·  Junaid World  ·  "
            f"Import Purchase Report — {_esc(firm_name)}  ·  "
            f"{datetime.now().strftime('%d %b %Y')}"
        )
        # Unescape for canvas (it doesn't use XML)
        footer_text = footer_text.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
        canvas.drawCentredString(pw / 2, footer_y, footer_text)

        # Confidentiality notice (right-aligned)
        canvas.setFont('Helvetica-Oblique', 5.5)
        canvas.setFillColor(HexColor('#CBD5E1'))
        canvas.drawRightString(pw - margin_h, footer_y, 'Confidential')

        canvas.restoreState()

    # ─────────────────────────────────────────────────────────────
    # BUILD & RETURN
    # ─────────────────────────────────────────────────────────────

    doc.build(elements, onFirstPage=_draw_page_chrome, onLaterPages=_draw_page_chrome)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    safe_filename = re.sub(r'[<>:"/\\|?*]', '_', firm_name)
    response['Content-Disposition'] = (
        f'attachment; filename="{safe_filename}_track_'
        f'{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    return response


@login_required
@admin_required
def export_consolidated_excel(request):
    """Export consolidated view to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from collections import defaultdict
    from datetime import datetime
    from io import BytesIO
    import re
    import math

    firm_name = request.GET.get('firm')
    search_raw = request.GET.get('search', '')
    search_tokens = parse_consolidated_search_tokens(search_raw)

    if not firm_name:
        return HttpResponse("Firm name required", status=400)

    quotation_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status__in=['CONFIRMED', 'COMPLETED']
    ).select_related('item', 'quotation').prefetch_related('releases').order_by('item__item_code')

    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'

    consolidated_data = {}
    all_dates = set()

    for q_item in quotation_items:
        on_the_way = q_item.releases.filter(is_received=False).aggregate(total=Sum('quantity_released'))['total'] or 0
        pending_at_factory = max(0, q_item.balance_to_release)
        is_fully_received = (on_the_way == 0 and pending_at_factory == 0)
        item = q_item.item
        key = (item.item_code, item.item_description, item.id)

        if not item_matches_consolidated_tokens(item.item_code, item.item_description, search_tokens):
            continue

        if key not in consolidated_data:
            consolidated_data[key] = {
                'dates': defaultdict(int),
                'on_the_way': 0,
                'pending_at_factory': 0,
                'total_qty': 0,
                'stock': item.item_stock or 0,
                'sold_stock': item.total_qty or 0 if is_admin else None,
                'reorder_qty': item.reorder_qty or 0,
                'is_fully_received': True,
            }

        if not is_fully_received:
            consolidated_data[key]['is_fully_received'] = False

        for release in q_item.releases.filter(is_received=False):
            if release.expected_arrival_date:
                date_str = release.expected_arrival_date.strftime('%b %d %Y')
                consolidated_data[key]['dates'][date_str] += release.quantity_released
                all_dates.add(date_str)

        consolidated_data[key]['on_the_way'] += on_the_way
        consolidated_data[key]['pending_at_factory'] += pending_at_factory
        consolidated_data[key]['total_qty'] += (on_the_way + pending_at_factory)

    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%b %d %Y'))

    local_map = fetch_local_open_qty_map()

    wb = openpyxl.Workbook()
    ws = wb.active
    sanitized_title = re.sub(r'[\/\\\?\*\[\]]', '_', firm_name[:31])
    ws.title = f"{sanitized_title} Consolidated"[:31]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Item Code', 'Item Name'] + sorted_dates + ['In Transit', 'To Be Released', 'Total Qty', 'Local Open Qty', 'Import + Local', 'Stock']
    if is_admin:
        headers += ['Sold Stock', 'Reorder Qty']
    else:
        headers += ['Reorder Qty']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def safe_val(val):
        if val is None:
            return ''
        if isinstance(val, (int, float)) and (math.isnan(val) or math.isinf(val)):
            return ''
        return val

    row_keys = ordered_consolidated_keys(list(consolidated_data.keys()), search_tokens)
    row_num = 2
    for key in row_keys:
        item_code, item_description, item_id = key
        data = consolidated_data[key]
        col = 1
        ws.cell(row=row_num, column=col, value=safe_val(item_code)).border = thin_border
        col += 1
        c = ws.cell(row=row_num, column=col, value=safe_val(item_description))
        c.border = thin_border
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        col += 1

        for date in sorted_dates:
            qty = data['dates'].get(date, 0)
            ws.cell(row=row_num, column=col, value=qty if qty > 0 else '-').border = thin_border
            col += 1

        for qty_field in ['on_the_way', 'pending_at_factory', 'total_qty']:
            v = data[qty_field]
            ws.cell(row=row_num, column=col, value=v if v > 0 else '-').border = thin_border
            col += 1

        local_open_qty = local_map.get(str(item_code).strip(), 0)
        import_plus_local = data['total_qty'] + local_open_qty
        ws.cell(row=row_num, column=col, value=local_open_qty if local_open_qty > 0 else '-').border = thin_border
        col += 1
        ws.cell(row=row_num, column=col, value=import_plus_local if import_plus_local > 0 else '-').border = thin_border
        col += 1

        ws.cell(row=row_num, column=col, value=safe_val(data['stock'])).border = thin_border
        col += 1
        if is_admin:
            ws.cell(row=row_num, column=col, value=safe_val(data['sold_stock']) if data['sold_stock'] is not None else '-').border = thin_border
            col += 1
        ws.cell(row=row_num, column=col, value=safe_val(data['reorder_qty'])).border = thin_border
        row_num += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_filename = re.sub(r'[<>:"/\\|?*]', '_', firm_name)
    response['Content-Disposition'] = f'attachment; filename="{safe_filename}_consolidated_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


@login_required
@admin_required
def export_consolidated_pdf(request):
    """
    Export consolidated import purchase view to PDF.
    Self-contained module — all helpers are local, no cross-app PDF dependencies.
    """
    from collections import defaultdict
    from datetime import datetime
    from io import BytesIO
    import os
    import re

    import requests
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.lib.units import inch, mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether,
    )

    # ─────────────────────────────────────────────────────────────
    # DESIGN TOKENS
    # ─────────────────────────────────────────────────────────────

    CLR_NAVY         = HexColor('#0F172A')
    CLR_NAVY_MED     = HexColor('#1E293B')
    CLR_SLATE        = HexColor('#334155')
    CLR_SLATE_LIGHT  = HexColor('#64748B')
    CLR_SLATE_MUTED  = HexColor('#94A3B8')
    CLR_ACCENT       = HexColor('#0EA5E9')
    CLR_ACCENT_DARK  = HexColor('#0284C7')
    CLR_BORDER       = HexColor('#CBD5E1')
    CLR_BORDER_HEAVY = HexColor('#94A3B8')
    CLR_ROW_ALT      = HexColor('#F8FAFC')
    CLR_ROW_WHITE    = colors.white
    CLR_HEADER_BG    = HexColor('#0F172A')
    CLR_HEADER_FG    = colors.white
    CLR_SECTION_BG   = HexColor('#F1F5F9')
    CLR_TOTAL_BG     = HexColor('#EFF6FF')
    CLR_HIGHLIGHT    = HexColor('#FEF3C7')    # Amber tint for non-zero dates
    CLR_ZERO_TEXT    = HexColor('#CBD5E1')     # Very muted for dashes
    CLR_WHITE        = colors.white
    CLR_SUCCESS      = HexColor('#059669')
    CLR_WARNING      = HexColor('#D97706')
    CLR_DANGER       = HexColor('#DC2626')

    FONT_TITLE       = 15
    FONT_SUBTITLE    = 9
    FONT_SECTION     = 10
    FONT_TH          = 6.5
    FONT_TD          = 6.5
    FONT_TD_SM       = 6
    FONT_META        = 8
    FONT_FOOTER      = 6
    FONT_BADGE       = 7
    FONT_KPI         = 9

    PAD_H            = 4
    PAD_V_HDR        = 6
    PAD_V_DATA       = 4
    SP_SECTION       = 12
    SP_INNER         = 5

    # ─────────────────────────────────────────────────────────────
    # LOCAL STYLES
    # ─────────────────────────────────────────────────────────────

    _base = getSampleStyleSheet()['Normal']

    def _ps(name, **kw):
        parent = kw.pop('parent', _base)
        return ParagraphStyle(name, parent=parent, **kw)

    S_TITLE = _ps('CTitle',
        fontName='Helvetica-Bold', fontSize=FONT_TITLE,
        textColor=CLR_NAVY, leading=FONT_TITLE + 4, alignment=TA_LEFT,
    )
    S_TITLE_R = _ps('CTitleR',
        fontName='Helvetica-Bold', fontSize=FONT_TITLE,
        textColor=CLR_NAVY, leading=FONT_TITLE + 4, alignment=TA_RIGHT,
    )
    S_SUBTITLE_R = _ps('CSubR',
        fontName='Helvetica', fontSize=FONT_SUBTITLE,
        textColor=CLR_SLATE_LIGHT, leading=FONT_SUBTITLE + 3, alignment=TA_RIGHT,
    )
    S_FIRM = _ps('CFirm',
        fontName='Helvetica', fontSize=11,
        textColor=CLR_ACCENT_DARK, leading=14, alignment=TA_RIGHT,
    )
    S_SECTION = _ps('CSec',
        fontName='Helvetica-Bold', fontSize=FONT_SECTION,
        textColor=CLR_NAVY, leading=FONT_SECTION + 3,
    )
    S_TH = _ps('CTH',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 2,
    )
    S_TH_C = _ps('CTHC',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 2, alignment=TA_CENTER,
    )
    S_TH_R = _ps('CTHR',
        fontName='Helvetica-Bold', fontSize=FONT_TH,
        textColor=CLR_WHITE, leading=FONT_TH + 2, alignment=TA_RIGHT,
    )
    S_TD = _ps('CTD',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2, wordWrap='CJK',
    )
    S_TD_C = _ps('CTDC',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2, alignment=TA_CENTER,
    )
    S_TD_R = _ps('CTDR',
        fontName='Helvetica', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2, alignment=TA_RIGHT,
    )
    S_TD_BOLD = _ps('CTDB',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2,
    )
    S_TD_BOLD_R = _ps('CTDBR',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2, alignment=TA_RIGHT,
    )
    S_TD_BOLD_C = _ps('CTDBC',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_SLATE, leading=FONT_TD + 2, alignment=TA_CENTER,
    )
    S_DASH = _ps('CDash',
        fontName='Helvetica', fontSize=FONT_TD_SM,
        textColor=CLR_ZERO_TEXT, leading=FONT_TD_SM + 2, alignment=TA_CENTER,
    )
    S_QTY_HIGHLIGHT = _ps('CQtyHi',
        fontName='Helvetica-Bold', fontSize=FONT_TD,
        textColor=CLR_ACCENT_DARK, leading=FONT_TD + 2, alignment=TA_CENTER,
    )
    S_ITEM_DESC = _ps('CItemDesc',
        fontName='Helvetica', fontSize=FONT_TD_SM,
        textColor=CLR_SLATE, leading=FONT_TD_SM + 2,
        wordWrap='CJK', leftIndent=2, rightIndent=2,
    )
    S_META = _ps('CMeta',
        fontName='Helvetica', fontSize=FONT_META,
        textColor=CLR_SLATE_LIGHT, leading=FONT_META + 3,
    )
    S_KPI_VAL = _ps('CKpiV',
        fontName='Helvetica-Bold', fontSize=FONT_KPI,
        textColor=CLR_NAVY, leading=FONT_KPI + 3, alignment=TA_CENTER,
    )
    S_KPI_LABEL = _ps('CKpiL',
        fontName='Helvetica', fontSize=FONT_FOOTER,
        textColor=CLR_SLATE_LIGHT, leading=FONT_FOOTER + 2, alignment=TA_CENTER,
    )
    S_EMPTY = _ps('CEmpty',
        fontName='Helvetica-Oblique', fontSize=FONT_TD,
        textColor=CLR_SLATE_LIGHT, leading=FONT_TD + 2, alignment=TA_CENTER,
    )

    # ─────────────────────────────────────────────────────────────
    # LOCAL HELPER FUNCTIONS
    # ─────────────────────────────────────────────────────────────

    def _esc(text):
        """Escape text for ReportLab XML."""
        s = str(text) if text is not None else ''
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def _p(text, style=None):
        return Paragraph(_esc(text), style or S_TD)

    def _p_num(value, style=None):
        try:
            v = float(value) if value else 0
            fmt = f"{v:,.0f}" if v == int(v) else f"{v:,.2f}"
        except (TypeError, ValueError):
            fmt = str(value or '0')
        return Paragraph(fmt, style or S_TD_R)

    def _p_dash():
        """Muted dash for zero/empty values."""
        return Paragraph('—', S_DASH)

    def _p_qty(value):
        """Quantity cell: highlighted if > 0, dash if 0."""
        if value and value > 0:
            return Paragraph(str(int(value)), S_QTY_HIGHLIGHT)
        return _p_dash()

    def _p_val_or_dash(value, style=None):
        """Show value if > 0, otherwise muted dash."""
        if value and value > 0:
            return _p_num(value, style or S_TD_BOLD_C)
        return _p_dash()

    def _load_logo():
        """Load logo bytes: local first, URL fallback."""
        local_paths = [
            os.path.join(settings.BASE_DIR, 'media', 'footer-logo.png'),
            os.path.join(settings.BASE_DIR, 'media', 'footer-logo1.png'),
            os.path.join(settings.BASE_DIR, 'static', 'images', 'footer-logo.png'),
        ]
        for path in local_paths:
            if os.path.exists(path):
                try:
                    with open(path, 'rb') as f:
                        return f.read()
                except Exception:
                    continue
        try:
            url = 'https://junaidworld.com/wp-content/uploads/2023/09/footer-logo.png.webp'
            r = requests.get(url, timeout=5)
            if r.status_code == 200 and r.content:
                return r.content
        except Exception:
            pass
        return None

    def _build_kpi_bar(items, page_width):
        """Build a horizontal KPI summary strip."""
        n = len(items)
        cell_w = page_width / n
        cells = []
        for label, value, clr in items:
            mini = Table([
                [Paragraph(f'<font color="{clr}"><b>{value}</b></font>', S_KPI_VAL)],
                [Paragraph(label, S_KPI_LABEL)],
            ], colWidths=[cell_w - 6])
            mini.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            cells.append(mini)

        bar = Table([cells], colWidths=[cell_w] * n)
        bar.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), CLR_SECTION_BG),
            ('BOX', (0, 0), (-1, -1), 0.5, CLR_BORDER),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, CLR_BORDER),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        return bar

    def _build_data_table(headers, data_rows, col_widths, date_col_start=None, date_col_end=None):
        """
        Build a professional data table.
        headers: list of (text, style) tuples
        date_col_start/end: column range for date columns (for special formatting)
        """
        header_row = [Paragraph(_esc(h[0]), h[1]) for h in headers]
        all_rows = [header_row] + data_rows

        tbl = Table(all_rows, colWidths=col_widths, repeatRows=1)

        num_rows = len(all_rows)
        cmds = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), CLR_HEADER_BG),
            ('TEXTCOLOR', (0, 0), (-1, 0), CLR_HEADER_FG),

            # Outer border
            ('BOX', (0, 0), (-1, -1), 0.75, CLR_BORDER_HEAVY),
            ('LINEBELOW', (0, 0), (-1, 0), 1, CLR_BORDER_HEAVY),

            # Padding — header
            ('TOPPADDING', (0, 0), (-1, 0), PAD_V_HDR),
            ('BOTTOMPADDING', (0, 0), (-1, 0), PAD_V_HDR),

            # Padding — data
            ('TOPPADDING', (0, 1), (-1, -1), PAD_V_DATA),
            ('BOTTOMPADDING', (0, 1), (-1, -1), PAD_V_DATA),
            ('LEFTPADDING', (0, 0), (-1, -1), PAD_H),
            ('RIGHTPADDING', (0, 0), (-1, -1), PAD_H),

            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Zebra striping + subtle row lines
        for i in range(1, num_rows):
            bg = CLR_ROW_ALT if i % 2 == 0 else CLR_ROW_WHITE
            cmds.append(('BACKGROUND', (0, i), (-1, i), bg))
            if i < num_rows - 1:
                cmds.append(('LINEBELOW', (0, i), (-1, i), 0.25, CLR_BORDER))

        # Vertical separator between date columns and summary columns
        if date_col_end is not None:
            cmds.append(('LINEAFTER', (date_col_end, 0), (date_col_end, -1), 1, CLR_BORDER_HEAVY))

        # Vertical separator between item description and date columns
        if date_col_start is not None:
            cmds.append(('LINEAFTER', (date_col_start - 1, 0), (date_col_start - 1, -1), 1, CLR_BORDER_HEAVY))

        # Last row bottom
        if num_rows > 1:
            cmds.append(('LINEBELOW', (0, -1), (-1, -1), 0.5, CLR_BORDER_HEAVY))

        tbl.setStyle(TableStyle(cmds))
        return tbl

    # ─────────────────────────────────────────────────────────────
    # BUSINESS LOGIC (unchanged)
    # ─────────────────────────────────────────────────────────────

    firm_name = request.GET.get('firm')
    search_raw = request.GET.get('search', '')
    search_tokens = parse_consolidated_search_tokens(search_raw)

    if not firm_name:
        return HttpResponse("Firm name required", status=400)

    quotation_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status__in=['CONFIRMED', 'COMPLETED']
    ).select_related('item', 'quotation').prefetch_related('releases').order_by('item__item_code')

    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'

    consolidated_data = {}
    all_dates = set()

    for q_item in quotation_items:
        on_the_way = q_item.releases.filter(is_received=False).aggregate(
            total=Sum('quantity_released')
        )['total'] or 0
        pending_at_factory = max(0, q_item.balance_to_release)
        is_fully_received = (on_the_way == 0 and pending_at_factory == 0)
        item = q_item.item
        key = (item.item_code, item.item_description, item.id)

        if not item_matches_consolidated_tokens(item.item_code, item.item_description, search_tokens):
            continue

        if key not in consolidated_data:
            consolidated_data[key] = {
                'dates': defaultdict(int),
                'on_the_way': 0,
                'pending_at_factory': 0,
                'total_qty': 0,
                'stock': item.item_stock or 0,
                'sold_stock': item.total_qty or 0 if is_admin else None,
                'reorder_qty': item.reorder_qty or 0,
                'is_fully_received': True,
            }

        if not is_fully_received:
            consolidated_data[key]['is_fully_received'] = False

        releases_on_way = q_item.releases.filter(is_received=False)
        for release in releases_on_way:
            if release.expected_arrival_date:
                date_str = release.expected_arrival_date.strftime('%b %d %Y')
                consolidated_data[key]['dates'][date_str] += release.quantity_released
                all_dates.add(date_str)

        consolidated_data[key]['on_the_way'] += on_the_way
        consolidated_data[key]['pending_at_factory'] += pending_at_factory
        consolidated_data[key]['total_qty'] += (on_the_way + pending_at_factory)

    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%b %d %Y'))

    local_map = fetch_local_open_qty_map()

    # Compute totals for KPI bar
    total_items = len(consolidated_data)
    total_transit = sum(d['on_the_way'] for d in consolidated_data.values())
    total_pending = sum(d['pending_at_factory'] for d in consolidated_data.values())
    total_all = sum(d['total_qty'] for d in consolidated_data.values())

    # ─────────────────────────────────────────────────────────────
    # PDF DOCUMENT SETUP
    # ─────────────────────────────────────────────────────────────

    logo_bytes = _load_logo()
    buffer = BytesIO()
    pw, ph = landscape(A4)
    margin_h = 0.35 * inch
    margin_top = 0.65 * inch
    margin_bot = 0.5 * inch
    usable_w = pw - 2 * margin_h

    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=margin_h, rightMargin=margin_h,
        topMargin=margin_top, bottomMargin=margin_bot,
    )

    elements = []

    # ─────────────────────────────────────────────────────────────
    # 1. DOCUMENT HEADER
    # ─────────────────────────────────────────────────────────────

    title_block = Paragraph('<b>Consolidated Import Report</b>', S_TITLE_R)
    firm_block = Paragraph(_esc(firm_name), S_FIRM)
    date_block = Paragraph(
        f"Generated: {datetime.now().strftime('%d %B %Y  •  %H:%M')}",
        S_SUBTITLE_R,
    )

    right_stack = Table(
        [[title_block], [firm_block], [date_block]],
        colWidths=[usable_w - 2.0 * inch],
    )
    right_stack.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))

    if logo_bytes:
        header_row = [['', right_stack]]
        header_widths = [2.0 * inch, usable_w - 2.0 * inch]
    else:
        brand = Paragraph('<b>JUNAID</b>', _ps('CBrand',
            fontName='Helvetica-Bold', fontSize=18, textColor=CLR_NAVY,
        ))
        header_row = [[brand, right_stack]]
        header_widths = [2.0 * inch, usable_w - 2.0 * inch]

    header_table = Table(header_row, colWidths=header_widths)
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 4))

    # Accent divider
    divider = Table([['']], colWidths=[usable_w])
    divider.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, CLR_ACCENT),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(divider)
    elements.append(Spacer(1, 6))

    # ─────────────────────────────────────────────────────────────
    # 2. KPI SUMMARY BAR
    # ─────────────────────────────────────────────────────────────

    kpi_items = [
        ('Total Items', str(total_items), '#0F172A'),
        ('In Transit', f"{total_transit:,}", '#2563EB'),
        ('At Factory', f"{total_pending:,}", '#D97706'),
        ('Total Qty', f"{total_all:,}", '#0F172A'),
        ('Arrival Dates', str(len(sorted_dates)), '#0EA5E9'),
    ]
    elements.append(_build_kpi_bar(kpi_items, usable_w))
    elements.append(Spacer(1, 4))

    # Search filter note
    if search_tokens:
        filter_label = _esc(search_raw.strip().replace('\n', ', ')[:200])
        if len(search_raw.strip()) > 200:
            filter_label += '…'
        elements.append(Paragraph(
            f'<font color="#64748B">Filter applied: "{filter_label}"</font>',
            S_META,
        ))
        elements.append(Spacer(1, 4))

    elements.append(Spacer(1, SP_SECTION - 4))

    # ─────────────────────────────────────────────────────────────
    # 3. SECTION HEADER
    # ─────────────────────────────────────────────────────────────

    accent_w = 4
    badge_w = 0.35 * inch
    content_w = usable_w - accent_w - badge_w - 1.0 * inch

    sec_badge = Paragraph('1', _ps('SecBadge',
        fontName='Helvetica-Bold', fontSize=9,
        textColor=CLR_WHITE, alignment=TA_CENTER,
    ))
    sec_title = Paragraph('CONSOLIDATED STOCK & ARRIVAL MATRIX', S_SECTION)

    section_bar = Table(
        [['', sec_badge, sec_title]],
        colWidths=[accent_w, badge_w, content_w],
    )
    section_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), CLR_ACCENT),
        ('BACKGROUND', (1, 0), (1, 0), CLR_ACCENT),
        ('BACKGROUND', (2, 0), (2, 0), CLR_SECTION_BG),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (2, 0), (2, 0), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, CLR_BORDER),
    ]))
    elements.append(section_bar)
    elements.append(Spacer(1, SP_INNER))

    # ─────────────────────────────────────────────────────────────
    # 4. CONSOLIDATED DATA TABLE
    # ─────────────────────────────────────────────────────────────

    # Build headers
    headers = [
        ('Code', S_TH),
        ('Item Description', S_TH),
    ]

    # Date columns — show abbreviated (e.g., "Mar 15")
    for d in sorted_dates:
        # Parse and reformat for shorter header
        try:
            dt = datetime.strptime(d, '%b %d %Y')
            short = dt.strftime('%d %b')
        except Exception:
            short = d[:6]
        headers.append((short, S_TH_C))

    date_col_start = 2
    date_col_end = 2 + len(sorted_dates) - 1

    headers.extend([
        ('Transit', S_TH_C),
        ('Pending', S_TH_C),
        ('Total', S_TH_C),
        ('Local', S_TH_C),
        ('Imp+Loc', S_TH_C),
        ('Stock', S_TH_C),
    ])

    if is_admin:
        headers.extend([('Sold', S_TH_C), ('Reorder', S_TH_C)])
    else:
        headers.append(('Reorder', S_TH_C))

    # Calculate column widths
    date_col_w = 38
    code_w = 52
    transit_w = 38
    pending_w = 38
    total_w = 36
    local_w = 34
    import_local_w = 38
    stock_w = 36
    sold_w = 34
    reorder_w = 38

    fixed_w = code_w + transit_w + pending_w + total_w + local_w + import_local_w + stock_w + reorder_w
    if is_admin:
        fixed_w += sold_w
    date_total_w = len(sorted_dates) * date_col_w
    desc_w = max(150, usable_w - fixed_w - date_total_w)

    col_widths = [code_w, desc_w]
    col_widths += [date_col_w] * len(sorted_dates)
    col_widths += [transit_w, pending_w, total_w, local_w, import_local_w, stock_w]
    if is_admin:
        col_widths += [sold_w, reorder_w]
    else:
        col_widths += [reorder_w]

    # Build data rows
    data_rows = []
    row_keys_pdf = ordered_consolidated_keys(list(consolidated_data.keys()), search_tokens)
    for key in row_keys_pdf:
        item_code, item_description, item_id = key
        data = consolidated_data[key]
        desc_para = Paragraph(_esc(item_description), S_ITEM_DESC)

        row = [
            _p(item_code, S_TD_BOLD),
            desc_para,
        ]

        if data.get('is_fully_received', False):
            # All received — show dashes for dates and summary
            for _ in sorted_dates:
                row.append(_p_dash())
            row.extend([_p_dash(), _p_dash(), _p_dash()])
            local_open_qty = local_map.get(str(item_code).strip(), 0)
            row.append(_p_num(local_open_qty, S_TD_BOLD_C) if local_open_qty > 0 else _p_dash())
            row.append(_p_num(local_open_qty, S_TD_BOLD_C) if local_open_qty > 0 else _p_dash())
            row.append(_p_num(data['stock'], S_TD_BOLD_C))
        else:
            # Active item — show date quantities and summary
            for date in sorted_dates:
                qty = data['dates'].get(date, 0)
                row.append(_p_qty(qty))
            row.append(_p_val_or_dash(data['on_the_way'], S_TD_BOLD_C))
            row.append(_p_val_or_dash(data['pending_at_factory'], S_TD_BOLD_C))
            row.append(_p_val_or_dash(data['total_qty'], S_TD_BOLD_C))
            local_open_qty = local_map.get(str(item_code).strip(), 0)
            import_plus_local = data['total_qty'] + local_open_qty
            row.append(_p_val_or_dash(local_open_qty, S_TD_BOLD_C))
            row.append(_p_val_or_dash(import_plus_local, S_TD_BOLD_C))
            row.append(_p_num(data['stock'], S_TD_BOLD_C))

        if is_admin:
            sold = data['sold_stock']
            row.append(Paragraph(str(sold) if sold is not None else '—', S_TD_C))
        row.append(_p_num(data['reorder_qty'], S_TD_C))

        data_rows.append(row)

    # Empty state
    if not data_rows:
        empty_row = [Paragraph('', S_TD)] * len(headers)
        empty_row[0] = Paragraph('No data found for this firm', S_EMPTY)
        data_rows.append(empty_row)

    # Build table
    main_table = _build_data_table(
        headers, data_rows, col_widths,
        date_col_start=date_col_start if sorted_dates else None,
        date_col_end=date_col_end if sorted_dates else None,
    )
    elements.append(main_table)

    # ─────────────────────────────────────────────────────────────
    # 5. LEGEND / FOOTNOTES
    # ─────────────────────────────────────────────────────────────

    elements.append(Spacer(1, 8))

    legend_items = [
        [
            Paragraph(
                '<font color="#0284C7"><b>●</b></font> '
                '<font color="#64748B" size="6">Highlighted numbers = quantities arriving on that date</font>',
                S_META,
            ),
            Paragraph(
                '<font color="#CBD5E1"><b>—</b></font> '
                '<font color="#64748B" size="6">Dash = zero / not applicable</font>',
                S_META,
            ),
            Paragraph(
                f'<font color="#64748B" size="6">Showing {total_items} items  •  '
                f'{len(sorted_dates)} arrival dates</font>',
                _ps('LegendR', parent=S_META, alignment=TA_RIGHT),
            ),
        ]
    ]
    legend_tbl = Table(legend_items, colWidths=[usable_w * 0.35, usable_w * 0.30, usable_w * 0.35])
    legend_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEABOVE', (0, 0), (-1, -1), 0.5, CLR_BORDER),
    ]))
    elements.append(legend_tbl)

    # ─────────────────────────────────────────────────────────────
    # PAGE HEADER & FOOTER CALLBACK
    # ─────────────────────────────────────────────────────────────

    def _draw_page_chrome(canvas, doc):
        canvas.saveState()

        # Logo in header
        if logo_bytes:
            try:
                img_reader = ImageReader(BytesIO(logo_bytes))
                canvas.drawImage(
                    img_reader,
                    margin_h, ph - margin_top + 6,
                    width=1.6 * inch, height=0.42 * inch,
                    preserveAspectRatio=True, mask='auto',
                )
            except Exception:
                pass

        # Footer
        footer_y = margin_bot - 18

        # Accent line
        canvas.setStrokeColor(CLR_ACCENT)
        canvas.setLineWidth(0.75)
        canvas.line(margin_h, footer_y + 12, pw - margin_h, footer_y + 12)

        # Footer text
        canvas.setFont('Helvetica', FONT_FOOTER)
        canvas.setFillColor(CLR_SLATE_LIGHT)
        raw_firm = firm_name.replace('&', '&').replace('<', '<').replace('>', '>')
        footer_text = (
            f"Page {doc.page}  ·  Junaid World  ·  "
            f"Consolidated Import Report — {raw_firm}  ·  "
            f"{datetime.now().strftime('%d %b %Y')}"
        )
        canvas.drawCentredString(pw / 2, footer_y, footer_text)

        # Confidentiality
        canvas.setFont('Helvetica-Oblique', 5.5)
        canvas.setFillColor(CLR_SLATE_MUTED)
        canvas.drawRightString(pw - margin_h, footer_y, 'Confidential')

        canvas.restoreState()

    # ─────────────────────────────────────────────────────────────
    # BUILD & RETURN
    # ─────────────────────────────────────────────────────────────

    doc.build(elements, onFirstPage=_draw_page_chrome, onLaterPages=_draw_page_chrome)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    safe_filename = re.sub(r'[<>:"/\\|?*]', '_', firm_name)
    response['Content-Disposition'] = (
        f'attachment; filename="{safe_filename}_consolidated_'
        f'{datetime.now().strftime("%Y%m%d")}.pdf"'
    )
    return response