import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum
from django.db.models.functions import Coalesce
from .models import ItemMaster, Quotation, QuotationItem, Release, Shipment, Manufacturer, LocalPurchaseItem
from .forms import UploadItemForm, QuotationForm, QuotationItemFormSet, ShipmentForm, ReleaseForm, ManufacturerForm, UploadManufacturerForm
from django.http import JsonResponse
import json
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .decorators import admin_required, sales_required
from django.core.management import call_command
from io import StringIO

@login_required
@admin_required
def dashboard(request):
    """
    Main dashboard view. Redirects based on user group (future).
    For now, nice landing page.
    """
    # Fetch recent quotations
    recent_quotations = Quotation.objects.all().order_by('-created_at')[:5]
    
    # Calculate stats
    total_items = ItemMaster.objects.count()
    pending_quotations = Quotation.objects.filter(status='DRAFT').count()
    
    # Incoming shipments (In Transit) - Release objects not received
    incoming_shipments = Release.objects.filter(is_received=False).count()

    context = {
        'recent_quotations': recent_quotations,
        'total_items': total_items,
        'pending_quotations': pending_quotations,
        'incoming_shipments': incoming_shipments,
    }
    return render(request, 'tracking/dashboard.html', context)

@login_required
@admin_required
def update_quotation_status(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)
    
    if request.method == 'POST':
        # Check if it's a form submission or JSON
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                new_status = data.get('status')
                if new_status in dict(Quotation.STATUS_CHOICES):
                    quotation.status = new_status
                    quotation.save()
                    return JsonResponse({'success': True})
                return JsonResponse({'success': False, 'error': 'Invalid status'})
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'error': 'Invalid JSON'})
        else:
            # Standard form submission (e.g. from the list page Save button)
            new_status = request.POST.get('status')
            if new_status in dict(Quotation.STATUS_CHOICES):
                quotation.status = new_status
                quotation.save()
                messages.success(request, f"Status for {quotation.reference_number} updated to {quotation.get_status_display()}.")
            else:
                messages.error(request, "Invalid status selected.")
            
            return redirect('quotation_list') # Redirect back to the list
            
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@never_cache
@login_required
@admin_required
def quotation_list(request):
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')
    
    quotes = Quotation.objects.all().order_by('-created_at')
    
    # helper for filter
    if status_filter != 'all':
        # exact match for status keys
        quotes = quotes.filter(status=status_filter.upper())

    # Search Logic
    if search_query:
        quotes = quotes.filter(
            models.Q(reference_number__icontains=search_query) |
            models.Q(supplier_name__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(quotes, 100) # 10 quotes per page
    page = request.GET.get('page')
    try:
        quotes_page = paginator.page(page)
    except PageNotAnInteger:
        quotes_page = paginator.page(1)
    except EmptyPage:
        quotes_page = paginator.page(paginator.num_pages)
    
    return render(request, 'tracking/quotation_list.html', {
        'quotes': quotes_page, 
        'filter': status_filter,
        'search': search_query
    })

@login_required
@admin_required
def release_item(request, pk):
    item = get_object_or_404(QuotationItem, pk=pk)
    
    # Check if there's anything available to release
    balance_to_release = item.balance_to_release
    if balance_to_release <= 0:
        messages.warning(request, f"No quantity available to release for {item.item.item_code}. Balance to release: {balance_to_release}")
        return redirect('quotation_detail', pk=item.quotation.pk)
    
    if request.method == 'POST':
        form = ReleaseForm(request.POST, quotation_item=item)
        if form.is_valid():
            quantity_released = form.cleaned_data['quantity_released']
            
            # Double-check validation (defense in depth)
            if quantity_released > balance_to_release:
                form.add_error('quantity_released', f"Cannot release more than available balance ({balance_to_release})")
            else:
                with transaction.atomic():
                    release = form.save(commit=False)
                    release.quotation_item = item
                    release.save()
                    
                    messages.success(request, f"Release created for {item.item.item_code}!")
                    return redirect('quotation_detail', pk=item.quotation.pk)
    else:
        # Pre-fill with available balance (but not more than available)
        initial_quantity = min(item.balance_to_release, item.balance_to_release) if item.balance_to_release > 0 else 0
        form = ReleaseForm(quotation_item=item, initial={'quantity_released': initial_quantity})

    return render(request, 'tracking/release_item.html', {
        'form': form,
        'item': item,
        'balance_to_release': balance_to_release
    })

@login_required
@admin_required
def receive_release(request, pk):
    # This view confirms that a Release (truck) has arrived
    release = get_object_or_404(Release, pk=pk)
    
    if request.method == 'POST':
        with transaction.atomic():
            release.is_received = True
            release.save()
            
            # Auto-create a Shipment record to log this receipt officially?
            # Or just mark release as received?
            # If we want to maintain the Shipment model as the source of "Stock", we should create one.
            Shipment.objects.create(
                quotation_item=release.quotation_item,
                quantity_received=release.quantity_released,
                received_date=pd.Timestamp.now().date(), # Default to today
                remarks=f"Auto-received from Release {release.container_info}"
            )
            
            # Check if quotation is fully received and update status automatically
            quotation = release.quotation_item.quotation
            if quotation.check_and_update_status():
                messages.success(request, f"Release received. Quotation {quotation.reference_number} status updated to COMPLETED - all items received!")
            else:
                messages.success(request, "Release marked as Received and Stock updated.")
            
            return redirect('quotation_detail', pk=release.quotation_item.quotation.pk)

    return render(request, 'tracking/receive_release_confirm.html', {'release': release})

@login_required
@admin_required
def create_quotation(request):
    if request.method == 'POST':
        print("DEBUG: POST keys received:", list(request.POST.keys()))
        print("DEBUG: POST items-0-item:", request.POST.get('items-0-item'))
        print("DEBUG: POST items-0-quantity_ordered:", request.POST.get('items-0-quantity_ordered'))
        
        form = QuotationForm(request.POST)
        formset = QuotationItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            print("DEBUG: Forms are valid")
            print(f"DEBUG: Total Forms: {formset.total_form_count()}")
            for i, f in enumerate(formset):
                print(f"DEBUG: Form {i} cleaned_data: {f.cleaned_data}")
                print(f"DEBUG: Form {i} has_changed: {f.has_changed()}")

            try:
                with transaction.atomic():
                    quotation = form.save()
                    quotation.created_by = request.user
                    quotation.save()
                    print(f"DEBUG: Quotation saved: {quotation.pk}")
                    
                    formset.instance = quotation
                    items = formset.save()
                    print(f"DEBUG: Items saved: {len(items)}")
                    
                    messages.success(request, f"Quotation {quotation.reference_number} created successfully!")
                    return redirect('dashboard')
            except Exception as e:
                print(f"Transaction Error: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error creating quotation: {str(e)}")
        else:
             print("DEBUG: Validation FAILED")
             print("Form Errors:", form.errors)
             print("Formset Errors:", formset.errors)
             print("POST Data (partial):", {k: v for k, v in request.POST.items() if 'item' in k})
             
             if formset.errors:
                 messages.error(request, f"Item Errors: {formset.errors}")
             messages.error(request, "Please correct the errors below.")
    else:
        form = QuotationForm()
        formset = QuotationItemFormSet()
        
    return render(request, 'tracking/create_quotation.html', {
        'form': form,
        'formset': formset
    })

@login_required
@admin_required
def quotation_detail(request, pk):
    quotation = get_object_or_404(Quotation, pk=pk)
    
    # Try to get supplier logo (use first supplier if multiple)
    supplier_logo = None
    try:
        from .models import Supplier
        # Get first supplier name from comma-separated list
        first_supplier_name = quotation.supplier_name.split(',')[0].strip() if quotation.supplier_name else None
        if first_supplier_name:
            supplier = Supplier.objects.filter(name=first_supplier_name).first()
            if supplier and supplier.logo:
                supplier_logo = supplier.logo
    except:
        pass
    
    return render(request, 'tracking/quotation_detail.html', {
        'quotation': quotation,
        'supplier_logo': supplier_logo,
    })

@login_required
@admin_required
def edit_quotation(request, pk):
    """Edit an existing quotation and its items."""
    quotation = get_object_or_404(Quotation, pk=pk)
    
    # Prevent editing confirmed/completed quotations
    if quotation.status in ['CONFIRMED', 'COMPLETED']:
        messages.warning(request, "Cannot edit a confirmed or completed quotation.")
        return redirect('quotation_detail', pk=pk)
    
    if request.method == 'POST':
        form = QuotationForm(request.POST, instance=quotation)
        formset = QuotationItemFormSet(request.POST, instance=quotation)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                    formset.save()
                    messages.success(request, f"Quotation {quotation.reference_number} updated successfully!")
                    return redirect('quotation_detail', pk=pk)
            except Exception as e:
                messages.error(request, f"Error updating quotation: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = QuotationForm(instance=quotation)
        formset = QuotationItemFormSet(instance=quotation)
    
    return render(request, 'tracking/edit_quotation.html', {
        'form': form,
        'formset': formset,
        'quotation': quotation,
    })

@login_required
@admin_required
def delete_quotation(request, pk):
    """Delete a quotation."""
    quotation = get_object_or_404(Quotation, pk=pk)
    
    # Prevent deleting confirmed/completed quotations
    if quotation.status in ['CONFIRMED', 'COMPLETED']:
        messages.warning(request, "Cannot delete a confirmed or completed quotation.")
        return redirect('quotation_detail', pk=pk)
    
    if request.method == 'POST':
        reference = quotation.reference_number
        quotation.delete()
        messages.success(request, f"Quotation {reference} deleted successfully!")
        return redirect('quotation_list')
    
    return render(request, 'tracking/delete_quotation.html', {
        'quotation': quotation,
    })

@login_required
@admin_required
def receive_item(request, pk):
    item = get_object_or_404(QuotationItem, pk=pk)
    
    # Get pending releases (In Transit) for this item to offer "Quick Receive"
    pending_releases = item.releases.filter(is_received=False)

    if request.method == 'POST':
        form = ShipmentForm(request.POST)
        if form.is_valid():
            shipment = form.save(commit=False)
            shipment.quotation_item = item
            shipment.save()
            messages.success(request, f"Received {shipment.quantity_received} of {item.item.item_code}")
            
            # Check if quotation is fully received and update status automatically
            quotation = item.quotation
            if quotation.check_and_update_status():
                messages.success(request, f"Quotation {quotation.reference_number} status updated to COMPLETED - all items received!")
            
            return redirect('quotation_detail', pk=item.quotation.pk)
    else:
        form = ShipmentForm()
    
    return render(request, 'tracking/receive_item.html', {
        'form': form,
        'item': item,
        'pending_releases': pending_releases
    })

@login_required
@admin_required
def upload_items(request):
    if request.method == 'POST':
        form = UploadItemForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                df = pd.read_excel(excel_file)
                
                # Basic validation: check required columns
                required_columns = ['Item Code', 'Item Description', 'Firm', 'Stock', 'UOM']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, f"Missing required columns. Expected: {', '.join(required_columns)}")
                    return render(request, 'tracking/upload_items.html', {'form': form})
                
                success_count = 0
                errors = []

                for index, row in df.iterrows():
                    try:
                        code = str(row['Item Code']).strip()
                        desc = str(row['Item Description']).strip()
                        firm = str(row['Firm']).strip()
                        stock = row['Stock'] if pd.notna(row['Stock']) else 0
                        uom = str(row['UOM']).strip()
                        
                        ItemMaster.objects.update_or_create(
                            item_code=code,
                            defaults={
                                'item_description': desc,
                                'item_firm': firm,
                                'item_stock': stock,
                                'uom': uom
                            }
                        )
                        success_count += 1
                    except Exception as e:
                        errors.append(f"Row {index+2}: {str(e)}")
                
                messages.success(request, f"Successfully processed {success_count} items.")
                if errors:
                    messages.warning(request, f"Encountered {len(errors)} errors. First few: {'; '.join(errors[:3])}")
                    
                return redirect('dashboard')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = UploadItemForm()
    
    return render(request, 'tracking/upload_items.html', {'form': form})

@never_cache
@login_required
@sales_required
def sales_dashboard(request):
    """
    Landing page for Sales. Select Firm.
    """
    # Get firms that have active QuotationItems (either pending release or in transit)
    # We want distinct firm names.
    firm_names = QuotationItem.objects.annotate(
        total_received=Coalesce(Sum('shipments__quantity_received'), 0)
    ).filter(
        models.Q(quotation__status='CONFIRMED') &
        models.Q(quantity_ordered__gt=models.F('total_received'))
    ).values_list('item__item_firm', flat=True).distinct()
    
    # Get supplier logos for these firms
    from .models import Supplier
    suppliers = {s.name: s for s in Supplier.objects.filter(name__in=firm_names)}
    
    # Build list of firms with their logos
    firms_with_logos = [
        {'name': firm, 'logo': suppliers.get(firm).logo if suppliers.get(firm) and suppliers.get(firm).logo else None}
        for firm in firm_names
    ]
    
    return render(request, 'tracking/sales_landing.html', {'firms': firms_with_logos})

@never_cache
@login_required
@sales_required
def sales_firm_track(request):
    firm_name = request.GET.get('firm')
    if not firm_name:
        return redirect('sales_dashboard')
        
    # 1. Incoming (On The Way)
    # Changed from grouping to flat list for Table view
    in_transit_releases = Release.objects.filter(
        quotation_item__item__item_firm=firm_name,
        is_received=False
    ).select_related(
        'quotation_item__item', 
        'quotation_item__quotation',
        'quotation_item__quotation__manufacturer'
    ).order_by('container_info', 'expected_arrival_date')
    
    # Add calculated amounts for admin display
    for release in in_transit_releases:
        release.transit_amount = release.quantity_released * release.quotation_item.rate

    # 2. Received (History) with Pagination
    received_queryset = Release.objects.filter(
        quotation_item__item__item_firm=firm_name,
        is_received=True
    ).select_related(
        'quotation_item__item', 
        'quotation_item__quotation',
        'quotation_item__quotation__manufacturer'
    ).prefetch_related(
        'quotation_item__shipments'
    ).order_by('-release_date')
    
    paginator = Paginator(received_queryset, 30) # Show 30 records per page
    page_number = request.GET.get('page')
    received_releases = paginator.get_page(page_number)
    
    # Add calculated amounts and received dates for admin display (after pagination)
    for release in received_releases:
        release.received_amount = release.quantity_released * release.quotation_item.rate
        # Get received date from shipment if available
        shipment = release.quotation_item.shipments.filter(
            quantity_received=release.quantity_released
        ).order_by('-received_date').first()
        if shipment:
            release.actual_received_date = shipment.received_date
        else:
            release.actual_received_date = release.release_date
    
    # 3. Pending (At Factory)
    all_firm_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status='CONFIRMED'
    ).select_related('item', 'quotation', 'quotation__manufacturer').order_by('expected_delivery_date')
    
    # Filter for items with balance > 0 and add calculated amounts for admin
    pending_items = []
    for item in all_firm_items:
        if item.balance_to_release > 0:
            # Add calculated amount for admin display
            item.balance_amount = item.balance_to_release * item.rate
            pending_items.append(item)
    
    # Get supplier logo
    supplier_logo = None
    try:
        from .models import Supplier
        supplier = Supplier.objects.filter(name=firm_name).first()
        if supplier and supplier.logo:
            supplier_logo = supplier.logo
    except:
        pass
    
    # Check if user is admin
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'
    
    # 4. Add stock information from ItemMaster model (updated by background process)
    # This is fast (database read) so we can show it to all users
    # Add stock information to in_transit_releases
    for release in in_transit_releases:
        release.current_stock = release.quotation_item.item.item_stock
        # Add sold stock (admin only)
        if is_admin:
            release.sold_stock = release.quotation_item.item.total_qty
    
    # Add stock information to received_releases
    for release in received_releases:
        release.current_stock = release.quotation_item.item.item_stock
        # Add sold stock (admin only)
        if is_admin:
            release.sold_stock = release.quotation_item.item.total_qty
    
    # Add stock information to pending_items
    for item in pending_items:
        item.current_stock = item.item.item_stock
        # Add sold stock (admin only)
        if is_admin:
            item.sold_stock = item.item.total_qty
    
    return render(request, 'tracking/sales_firm_track.html', {
        'firm': firm_name,
        'in_transit_releases': in_transit_releases, # Updated context variable
        'received_releases': received_releases,
        'pending_items': pending_items,
        'supplier_logo': supplier_logo,
        'is_admin': is_admin,
    })

@never_cache
@login_required
def get_items_by_firm(request):
    firm = request.GET.get('firm')
    if not firm:
        return JsonResponse({'items': []})
    
    items = ItemMaster.objects.filter(item_firm=firm).values('id', 'item_code', 'item_description', 'item_upvc').order_by('item_code')
    return JsonResponse({'items': list(items)})

@login_required
def consolidated_view(request):
    """
    Consolidated view showing all items for a firm with:
    - Expected arrival date columns (with quantities)
    - Total Qty On The Way (released but not received)
    - Total Qty Pending at Factory (not yet released)
    First shows firm selection, then the consolidated table.
    """
    from collections import defaultdict
    from datetime import datetime
    
    firm_name = request.GET.get('firm')
    
    # If no firm selected, show firm selection page
    if not firm_name:
        # Get firms that have active QuotationItems
        firm_names = QuotationItem.objects.filter(
            quotation__status__in=['CONFIRMED', 'COMPLETED']
        ).values_list('item__item_firm', flat=True).distinct().order_by('item__item_firm')
        
        # Get supplier logos
        from .models import Supplier
        suppliers = {s.name: s for s in Supplier.objects.filter(name__in=firm_names)}
        
        firms_with_logos = [
            {'name': firm, 'logo': suppliers.get(firm).logo if suppliers.get(firm) and suppliers.get(firm).logo else None}
            for firm in firm_names
        ]
        
        return render(request, 'tracking/consolidated_firm_select.html', {'firms': firms_with_logos})
    
    # Get all quotation items for this firm that have orders (including fully received)
    quotation_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status__in=['CONFIRMED', 'COMPLETED']
    ).select_related('item', 'quotation').prefetch_related('releases').order_by('item__item_code')
    
    # Check if user is admin
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'
    
    # Build consolidated data structure
    # Key: (item_code, item_description, item_id)
    # Value: {
    #   'dates': {date: quantity} - only for on the way items,
    #   'on_the_way': sum of releases where is_received=False,
    #   'pending_at_factory': balance_to_release,
    #   'total_qty': on_the_way + pending_at_factory,
    #   'stock': value,
    #   'sold_stock': value,
    #   'is_fully_received': boolean - True if all items received
    # }
    consolidated_data = {}
    all_dates = set()
    
    for q_item in quotation_items:
        # Calculate on the way (released but not received)
        on_the_way = q_item.releases.filter(is_received=False).aggregate(
            total=Sum('quantity_released')
        )['total'] or 0
        
        # Calculate pending at factory (balance_to_release) - ensure non-negative
        pending_at_factory = max(0, q_item.balance_to_release)
        
        # Check if fully received (no pending or in-transit)
        is_fully_received = (on_the_way == 0 and pending_at_factory == 0)
        
        item = q_item.item
        key = (item.item_code, item.item_description, item.id)
        
        if key not in consolidated_data:
            consolidated_data[key] = {
                'dates': defaultdict(int),
                'on_the_way': 0,
                'pending_at_factory': 0,
                'total_qty': 0,
                'stock': item.item_stock or 0,
                'sold_stock': item.total_qty or 0 if is_admin else None,
                'reorder_qty': item.reorder_qty or 0,
                'is_fully_received': True,  # Will be set to False if any order is not fully received
            }
        
        # If this order is not fully received, mark the item as not fully received
        if not is_fully_received:
            consolidated_data[key]['is_fully_received'] = False
        
        # Get only releases that are on the way (not received) grouped by expected_arrival_date
        releases_on_way = q_item.releases.filter(is_received=False)
        for release in releases_on_way:
            if release.expected_arrival_date:
                # Use shorter date format for compact display
                date_str = release.expected_arrival_date.strftime('%b %d %Y')
                consolidated_data[key]['dates'][date_str] += release.quantity_released
                all_dates.add(date_str)
        
        # Add quantities
        consolidated_data[key]['on_the_way'] += on_the_way
        consolidated_data[key]['pending_at_factory'] += pending_at_factory
        consolidated_data[key]['total_qty'] += (on_the_way + pending_at_factory) 
    
    # Sort dates chronologically
    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%b %d %Y'))
    
    # Convert to list for template
    table_data = []
    for (item_code, item_description, item_id), data in sorted(consolidated_data.items()):
        # Create a list of quantities matching the sorted_dates order
        date_qty_list = [data['dates'].get(date, 0) for date in sorted_dates]
        row = {
            'item_code': item_code,
            'item_description': item_description,
            'item_id': item_id,
            'on_the_way': data['on_the_way'],
            'pending_at_factory': data['pending_at_factory'],
            'total_qty': data['total_qty'],
            'stock': data['stock'],
            'sold_stock': data['sold_stock'],
            'reorder_qty': data['reorder_qty'],
            'date_quantities': date_qty_list,  # List matching sorted_dates order
            'is_fully_received': data['is_fully_received'],  # True if all orders for this item are received
        }
        table_data.append(row)
    
    # Get supplier logo
    supplier_logo = None
    try:
        from .models import Supplier
        supplier = Supplier.objects.filter(name=firm_name).first()
        if supplier and supplier.logo:
            supplier_logo = supplier.logo
    except:
        pass
    
    return render(request, 'tracking/consolidated_view.html', {
        'firm': firm_name,
        'table_data': table_data,
        'dates': sorted_dates,
        'supplier_logo': supplier_logo,
        'is_admin': is_admin,
    })

@login_required
@admin_required
def update_reorder_qty(request):
    """AJAX endpoint to update reorder quantity for an item"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST allowed'})
    
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id')
        reorder_qty = data.get('reorder_qty', 0)
        
        if not item_id:
            return JsonResponse({'success': False, 'error': 'Item ID required'})
        
        item = ItemMaster.objects.get(id=item_id)
        item.reorder_qty = int(reorder_qty) if reorder_qty else 0
        item.save(update_fields=['reorder_qty'])
        
        return JsonResponse({'success': True, 'reorder_qty': item.reorder_qty})
    except ItemMaster.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@admin_required
def reset_reorder_qty(request):
    """AJAX endpoint to reset all reorder quantities for a firm to 0"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST allowed'})
    
    try:
        data = json.loads(request.body)
        firm_name = data.get('firm')
        
        if not firm_name:
            return JsonResponse({'success': False, 'error': 'Firm name required'})
        
        # Reset all items for this firm
        updated_count = ItemMaster.objects.filter(item_firm=firm_name).update(reorder_qty=0)
        
        return JsonResponse({
            'success': True, 
            'message': f'Reset {updated_count} items to 0',
            'updated_count': updated_count
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@admin_required
def export_consolidated_excel(request):
    """Export consolidated view to Excel"""
    import openpyxl
    import openpyxl.utils
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from collections import defaultdict
    from datetime import datetime
    
    firm_name = request.GET.get('firm')
    search_query = request.GET.get('search', '').strip().lower()
    
    if not firm_name:
        return HttpResponse("Firm name required", status=400)
    
    # Get data (same logic as consolidated_view)
    quotation_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status__in=['CONFIRMED', 'COMPLETED']
    ).select_related('item', 'quotation').prefetch_related('releases').order_by('item__item_code')
    
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'
    
    consolidated_data = {}
    all_dates = set()
    
    for q_item in quotation_items:
        on_the_way = q_item.releases.filter(is_received=False).aggregate(total=Sum('quantity_released'))['total'] or 0
        pending_at_factory = max(0, q_item.balance_to_release)
        
        # Check if fully received (no pending or in-transit)
        is_fully_received = (on_the_way == 0 and pending_at_factory == 0)
        
        item = q_item.item
        key = (item.item_code, item.item_description, item.id)
        
        # Apply search filter
        if search_query:
            if search_query not in item.item_code.lower() and search_query not in item.item_description.lower():
                continue
        
        if key not in consolidated_data:
            consolidated_data[key] = {
                'dates': defaultdict(int),
                'on_the_way': 0,
                'pending_at_factory': 0,
                'total_qty': 0,
                'stock': item.item_stock or 0,
                'sold_stock': item.total_qty or 0 if is_admin else None,
                'reorder_qty': item.reorder_qty or 0,
                'is_fully_received': True,  # Will be set to False if any order is not fully received
            }
        
        # If this order is not fully received, mark the item as not fully received
        if not is_fully_received:
            consolidated_data[key]['is_fully_received'] = False
        
        releases_on_way = q_item.releases.filter(is_received=False)
        for release in releases_on_way:
            if release.expected_arrival_date:
                date_str = release.expected_arrival_date.strftime('%b %d %Y')
                consolidated_data[key]['dates'][date_str] += release.quantity_released
                all_dates.add(date_str)
        
        consolidated_data[key]['on_the_way'] += on_the_way
        consolidated_data[key]['pending_at_factory'] += pending_at_factory
        consolidated_data[key]['total_qty'] += (on_the_way + pending_at_factory)
    
    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%b %d %Y'))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{firm_name[:20]} Consolidated"
    
    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Item Code', 'Item Name'] + sorted_dates + ['In Transit', 'To Be Released', 'Total Qty', 'Stock']
    if is_admin:
        headers += ['Sold Stock', 'Reorder Qty']
    else:
        headers += ['Reorder Qty']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data rows
    row_num = 2
    for (item_code, item_description, item_id), data in sorted(consolidated_data.items()):
        col = 1
        ws.cell(row=row_num, column=col, value=item_code).border = thin_border
        col += 1
        # Item name cell with wrap text enabled
        item_name_cell = ws.cell(row=row_num, column=col, value=item_description)
        item_name_cell.border = thin_border
        item_name_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        col += 1
        
        # Show '-' for fully received items
        if data.get('is_fully_received', False):
            for date in sorted_dates:
                ws.cell(row=row_num, column=col, value='-').border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value='-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value='-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value='-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=data['stock']).border = thin_border
            col += 1
            if is_admin:
                ws.cell(row=row_num, column=col, value=data['sold_stock'] if data['sold_stock'] is not None else '-').border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value=data['reorder_qty']).border = thin_border
        else:
            for date in sorted_dates:
                qty = data['dates'].get(date, 0)
                ws.cell(row=row_num, column=col, value=qty if qty > 0 else '-').border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value=data['on_the_way'] if data['on_the_way'] > 0 else '-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=data['pending_at_factory'] if data['pending_at_factory'] > 0 else '-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=data['total_qty'] if data['total_qty'] > 0 else '-').border = thin_border
            col += 1
            ws.cell(row=row_num, column=col, value=data['stock']).border = thin_border
            col += 1
            if is_admin:
                ws.cell(row=row_num, column=col, value=data['sold_stock'] if data['sold_stock'] is not None else '-').border = thin_border
                col += 1
            ws.cell(row=row_num, column=col, value=data['reorder_qty']).border = thin_border
        
        row_num += 1
    
    # Adjust column widths - make item name column wider
    ws.column_dimensions['A'].width = 12  # Item Code
    ws.column_dimensions['B'].width = 60  # Item Name - increased from 40 to 60 for full names
    
    # Set widths for date columns
    date_start_col = 3  # Column C (after Code and Name)
    for i in range(len(sorted_dates)):
        col_letter = openpyxl.utils.get_column_letter(date_start_col + i)
        ws.column_dimensions[col_letter].width = 12
    
    # Set widths for remaining columns
    remaining_start = date_start_col + len(sorted_dates)
    remaining_cols = ['In Transit', 'To Be Released', 'Total Qty', 'Stock']
    if is_admin:
        remaining_cols += ['Sold Stock', 'Reorder Qty']
    else:
        remaining_cols += ['Reorder Qty']
    
    for i, col_name in enumerate(remaining_cols):
        col_letter = openpyxl.utils.get_column_letter(remaining_start + i)
        if col_name in ['In Transit', 'To Be Released', 'Total Qty']:
            ws.column_dimensions[col_letter].width = 12
        elif col_name in ['Stock', 'Sold Stock', 'Reorder Qty']:
            ws.column_dimensions[col_letter].width = 10
        else:
            ws.column_dimensions[col_letter].width = 12
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{firm_name}_consolidated_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_consolidated_pdf(request):
    """Export consolidated view to PDF"""
    from django.http import HttpResponse
    from collections import defaultdict
    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from io import BytesIO
    
    firm_name = request.GET.get('firm')
    search_query = request.GET.get('search', '').strip().lower()
    
    if not firm_name:
        return HttpResponse("Firm name required", status=400)
    
    # Get data (same logic as consolidated_view)
    quotation_items = QuotationItem.objects.filter(
        item__item_firm=firm_name,
        quotation__status__in=['CONFIRMED', 'COMPLETED']
    ).select_related('item', 'quotation').prefetch_related('releases').order_by('item__item_code')
    
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'ADMIN'
    
    consolidated_data = {}
    all_dates = set()
    
    for q_item in quotation_items:
        on_the_way = q_item.releases.filter(is_received=False).aggregate(total=Sum('quantity_released'))['total'] or 0
        pending_at_factory = max(0, q_item.balance_to_release)
        
        # Check if fully received (no pending or in-transit)
        is_fully_received = (on_the_way == 0 and pending_at_factory == 0)
        
        item = q_item.item
        key = (item.item_code, item.item_description, item.id)
        
        # Apply search filter
        if search_query:
            if search_query not in item.item_code.lower() and search_query not in item.item_description.lower():
                continue
        
        if key not in consolidated_data:
            consolidated_data[key] = {
                'dates': defaultdict(int),
                'on_the_way': 0,
                'pending_at_factory': 0,
                'total_qty': 0,
                'stock': item.item_stock or 0,
                'sold_stock': item.total_qty or 0 if is_admin else None,
                'reorder_qty': item.reorder_qty or 0,
                'is_fully_received': True,  # Will be set to False if any order is not fully received
            }
        
        # If this order is not fully received, mark the item as not fully received
        if not is_fully_received:
            consolidated_data[key]['is_fully_received'] = False
        
        releases_on_way = q_item.releases.filter(is_received=False)
        for release in releases_on_way:
            if release.expected_arrival_date:
                date_str = release.expected_arrival_date.strftime('%b %d %Y')
                consolidated_data[key]['dates'][date_str] += release.quantity_released
                all_dates.add(date_str)
        
        consolidated_data[key]['on_the_way'] += on_the_way
        consolidated_data[key]['pending_at_factory'] += pending_at_factory
        consolidated_data[key]['total_qty'] += (on_the_way + pending_at_factory)
    
    sorted_dates = sorted(all_dates, key=lambda x: datetime.strptime(x, '%b %d %Y'))
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=0.3*inch, rightMargin=0.3*inch)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, spaceAfter=12)
    
    elements.append(Paragraph(f"{firm_name} - Import Purchase Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Build table data with Paragraph for text wrapping
    # Create a style for item names that allows wrapping
    item_name_style = ParagraphStyle(
        'ItemName',
        parent=styles['Normal'],
        fontSize=7,
        leading=8,
        leftIndent=2,
        rightIndent=2,
        wordWrap='CJK'  # Allows word wrapping
    )
    
    headers = ['Code', 'Item Name'] + [d[:6] for d in sorted_dates] + ['Transit', 'Pending', 'Total', 'Stock']
    if is_admin:
        headers += ['Sold', 'Reorder']
    else:
        headers += ['Reorder']
    
    table_data = [headers]
    
    for (item_code, item_description, item_id), data in sorted(consolidated_data.items()):
        # Use Paragraph for item description to enable text wrapping
        # Escape special characters for XML/PDF
        item_desc_escaped = item_description.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        item_name_para = Paragraph(item_desc_escaped, item_name_style)
        
        # Show '-' for fully received items
        if data.get('is_fully_received', False):
            row = [item_code, item_name_para]
            for date in sorted_dates:
                row.append('-')
            row.extend(['-', '-', '-', data['stock']])
            if is_admin:
                row.append(data['sold_stock'] if data['sold_stock'] is not None else '-')
            row.append(data['reorder_qty'])
        else:
            row = [item_code, item_name_para]
            for date in sorted_dates:
                qty = data['dates'].get(date, 0)
                row.append(qty if qty > 0 else '-')
            row.extend([
                data['on_the_way'] if data['on_the_way'] > 0 else '-',
                data['pending_at_factory'] if data['pending_at_factory'] > 0 else '-',
                data['total_qty'] if data['total_qty'] > 0 else '-',
                data['stock']
            ])
            if is_admin:
                row.append(data['sold_stock'] if data['sold_stock'] is not None else '-')
            row.append(data['reorder_qty'])
        
        table_data.append(row)
    
    # Create table with wider column for item name
    # Calculate available width (landscape A4 width - margins)
    # landscape(A4) returns (height, width), so we use [1] for width
    available_width = landscape(A4)[1] - (0.3 * inch * 2)  # Subtract left and right margins
    date_col_width = 35
    fixed_cols_width = 55 + 250 + 40 + 40 + 35 + 35  # Code + Item Name + Transit + Pending + Total + Stock
    if is_admin:
        fixed_cols_width += 35 + 40  # Sold + Reorder
    else:
        fixed_cols_width += 40  # Reorder
    
    # Adjust item name width if we have many date columns
    item_name_width = 250
    if len(sorted_dates) > 0:
        date_cols_width = len(sorted_dates) * date_col_width
        remaining = available_width - fixed_cols_width - date_cols_width
        if remaining > 0:
            item_name_width += remaining
        elif remaining < 0:
            # Need to reduce item name width
            item_name_width = max(200, 250 + remaining)
    
    col_widths = [55, item_name_width] + [date_col_width] * len(sorted_dates) + [40, 40, 35, 35]
    if is_admin:
        col_widths += [35, 40]
    else:
        col_widths += [40]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),  # Item name left-aligned
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Top align for wrapped text
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{firm_name}_consolidated_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

@login_required
@admin_required
def upload_quotation_items_excel(request):
    """AJAX endpoint to upload Excel file with quotation items (itemcode, qty, rate)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'})
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})
    
    excel_file = request.FILES['excel_file']
    
    # Validate file extension
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls file'})
    
    try:
        # Read Excel file - skip empty rows at the start
        df = pd.read_excel(excel_file)
        
        # Drop completely empty rows (all NaN)
        df = df.dropna(how='all')
        
        # Reset index after dropping empty rows
        df = df.reset_index(drop=True)
        
        # Normalize column names (case-insensitive, handle spaces)
        df.columns = [str(col).strip().lower().replace(' ', '_') for col in df.columns]
        
        # Find column names (flexible matching)
        itemcode_col = None
        qty_col = None
        rate_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if itemcode_col is None and ('itemcode' in col_lower or 'item_code' in col_lower):
                itemcode_col = col
            if qty_col is None and ('qty' in col_lower or 'quantity' in col_lower):
                qty_col = col
            if rate_col is None and col_lower == 'rate':
                rate_col = col
        
        # Validate required columns
        if not itemcode_col:
            return JsonResponse({'success': False, 'error': 'Could not find itemcode column. Expected: itemcode, item_code, or Item Code'})
        if not qty_col:
            return JsonResponse({'success': False, 'error': 'Could not find quantity column. Expected: qty, quantity, or Quantity'})
        if not rate_col:
            return JsonResponse({'success': False, 'error': 'Could not find rate column. Expected: rate or Rate'})
        
        items = []
        warnings = []
        
        # Process each row
        for index, row in df.iterrows():
            # Get itemcode value - handle Excel float conversion (e.g., "300242.0" -> "300242")
            itemcode_raw = row[itemcode_col]
            
            # Skip if NaN or empty
            if pd.isna(itemcode_raw):
                warnings.append(f"Row {index + 2}: Skipped - empty itemcode")
                continue
            
            # Convert to string and handle float conversion
            # If it's a float like 300242.0, convert to int first, then string
            try:
                # Try to convert to float first (handles "300242.0")
                if isinstance(itemcode_raw, (int, float)):
                    # If it's a number, convert to int (removes .0) then string
                    itemcode = str(int(float(itemcode_raw)))
                else:
                    # It's already a string, but might be "300242.0"
                    itemcode_str = str(itemcode_raw).strip()
                    # Try to parse as float, then convert to int to remove .0
                    try:
                        itemcode = str(int(float(itemcode_str)))
                    except (ValueError, TypeError):
                        # If it's not a number, use as-is
                        itemcode = itemcode_str
            except (ValueError, TypeError):
                itemcode = str(itemcode_raw).strip()
            
            # Skip if empty after conversion
            if not itemcode or itemcode.lower() == 'nan':
                warnings.append(f"Row {index + 2}: Skipped - empty itemcode")
                continue
            
            # Skip if the itemcode value matches header column names (case-insensitive)
            header_variations = ['itemcode', 'item_code', 'item code', 'item', 'code']
            if itemcode.lower() in header_variations:
                # This is likely a header row that pandas didn't recognize, skip it
                continue
            
            # Try to find item in ItemMaster (case-insensitive match)
            try:
                item_master = ItemMaster.objects.get(item_code__iexact=itemcode)
            except ItemMaster.DoesNotExist:
                warnings.append(f"Row {index + 2}: Itemcode '{itemcode}' not found in database")
                continue
            except ItemMaster.MultipleObjectsReturned:
                # If multiple found, take first one
                item_master = ItemMaster.objects.filter(item_code__iexact=itemcode).first()
            
            # Get quantity (default to 0 if invalid)
            try:
                qty = int(float(row[qty_col])) if pd.notna(row[qty_col]) else 0
            except (ValueError, TypeError):
                qty = 0
                warnings.append(f"Row {index + 2}: Invalid quantity for '{itemcode}', using 0")
            
            # Get rate (default to 0.0 if invalid) and round to 2 decimal places
            try:
                rate = float(row[rate_col]) if pd.notna(row[rate_col]) else 0.0
                rate = round(rate, 2)  # Round to 2 decimal places
            except (ValueError, TypeError):
                rate = 0.0
                warnings.append(f"Row {index + 2}: Invalid rate for '{itemcode}', using 0.0")
            
            # Add to items list
            items.append({
                'item_id': item_master.id,
                'item_code': item_master.item_code,
                'item_description': item_master.item_description,
                'item_upvc': item_master.item_upvc or '',
                'quantity_ordered': qty,
                'rate': round(float(rate), 2)  # Ensure 2 decimal places
            })
        
        if not items:
            return JsonResponse({
                'success': False,
                'error': 'No valid items found in Excel file',
                'warnings': warnings
            })
        
        return JsonResponse({
            'success': True,
            'items': items,
            'warnings': warnings,
            'count': len(items)
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Excel upload error: {error_trace}")  # Log to console for debugging
        return JsonResponse({
            'success': False,
            'error': f'Error processing Excel file: {str(e)}',
            'traceback': error_trace if request.user.is_superuser else None  # Only show traceback to superusers
        })

# Manufacturer Management
@login_required
@admin_required
def manufacturer_list(request):
    manufacturers = Manufacturer.objects.all()
    return render(request, 'tracking/manufacturer_list.html', {
        'manufacturers': manufacturers
    })

@login_required
@admin_required
def manufacturer_create(request):
    if request.method == 'POST':
        form = ManufacturerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Manufacturer added successfully!")
            return redirect('manufacturer_list')
    else:
        form = ManufacturerForm()
    return render(request, 'tracking/manufacturer_form.html', {
        'form': form,
        'title': 'Add Manufacturer'
    })

@login_required
@admin_required
def manufacturer_edit(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    if request.method == 'POST':
        form = ManufacturerForm(request.POST, instance=manufacturer)
        if form.is_valid():
            form.save()
            messages.success(request, "Manufacturer updated successfully!")
            return redirect('manufacturer_list')
    else:
        form = ManufacturerForm(instance=manufacturer)
    return render(request, 'tracking/manufacturer_form.html', {
        'form': form,
        'title': 'Edit Manufacturer'
    })

@login_required
@admin_required
def manufacturer_delete(request, pk):
    manufacturer = get_object_or_404(Manufacturer, pk=pk)
    if request.method == 'POST':
        name = manufacturer.name
        manufacturer.delete()
        messages.success(request, f"Manufacturer {name} deleted successfully!")
        return redirect('manufacturer_list')
    return render(request, 'tracking/manufacturer_confirm_delete.html', {
        'manufacturer': manufacturer
    })

@login_required
@admin_required
def manufacturer_upload(request):
    if request.method == 'POST':
        form = UploadManufacturerForm(request.POST, request.FILES)
        if form.is_valid():
            import pandas as pd
            file = request.FILES['file']
            try:
                df = pd.read_excel(file)
                # Expecting a column 'Manufacturer' or 'Name'
                column_name = None
                for col in ['Manufacturer', 'Name', 'manufacturer', 'name']:
                    if col in df.columns:
                        column_name = col
                        break
                
                if not column_name:
                    messages.error(request, "Could not find a 'Manufacturer' or 'Name' column in Excel.")
                    return render(request, 'tracking/manufacturer_upload.html', {'form': form})
                
                count = 0
                for _, row in df.iterrows():
                    name = str(row[column_name]).strip()
                    if name and name != 'nan':
                        Manufacturer.objects.get_or_create(name=name)
                        count += 1
                
                messages.success(request, f"Successfully imported {count} manufacturers!")
                return redirect('manufacturer_list')
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    else:
        form = UploadManufacturerForm()
    return render(request, 'tracking/manufacturer_upload.html', {'form': form})

@login_required
@admin_required
def run_stock_import(request):
    out = StringIO()
    try:
        call_command('import_stock_api', stdout=out)
        output = out.getvalue()
        messages.success(request, output)
    except Exception as e:
        messages.error(request, f"Error importing stock data: {str(e)}")
    finally:
        out.close()
    return redirect('dashboard')

@login_required
@admin_required
def local_purchase_dashboard(request):
    """Dashboard to select a brand/sheet."""
    brands = LocalPurchaseItem.objects.values_list('brand', flat=True).distinct().order_by('brand')
    return render(request, 'tracking/local_purchase_dashboard.html', {'brands': brands})

from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.http import JsonResponse

@login_required
@admin_required
def local_purchase_list(request):
    """Table view for a selected brand with AJAX search, sort, and pagination."""
    brand = request.GET.get('brand')
    if not brand:
        return redirect('local_purchase_dashboard')
        
    # Start with base queryset
    items = LocalPurchaseItem.objects.filter(brand=brand)

    # 1. Search Filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        items = items.filter(
            Q(item_code__icontains=search_query) |
            Q(upc_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # 2. Quick Filter (migrated from JS to Backend)
    quick_filter = request.GET.get('filter', 'all')
    if quick_filter == 'critical':
        items = items.filter(stock_sufficiency_months__lt=1)
    elif quick_filter == 'required':
        # Assuming 'Stock Requirement' > 0 implies requirement exists
        items = items.filter(stock_requirement__gt=0) 
        # Note: Adjust field lookup based on your exact model type (int/str)
    elif quick_filter == 'high-value':
        items = items.filter(value__gt=5000)

    # 3. Sorting
    sort_by = request.GET.get('sort', 'item_code') # default sort
    direction = request.GET.get('direction', 'asc')
    
    # Whitelist allowable sort fields to prevent errors
    allowed_sort_fields = {
        'item_code': 'item_code',
        'upc_code': 'upc_code',
        'description': 'description',
        'current_stock_ras': 'current_stock_ras',
        'current_stock_dip': 'current_stock_dip',
        'sold_qty_2024': 'sold_qty_2024',
        'contg': 'contg',
        'trdg': 'trdg',
        'stores': 'stores',
        'total_sold_qty_2025': 'total_sold_qty_2025',
        'avg_15day_sales': 'avg_15day_sales',
        'stock_sufficiency_months': 'stock_sufficiency_months',
        'lpo_given': 'lpo_given',
        'open_so_qty': 'open_so_qty',
        'stock_reqt_calcn': 'stock_reqt_calcn',
        'stock_requirement': 'stock_requirement',
        'value': 'value',
        'cost': 'cost',
        'ho_per_lpo_qty': 'ho_per_lpo_qty',
        'stock_reqt_ras_stores': 'stock_reqt_ras_stores',
    }
    
    sort_field = allowed_sort_fields.get(sort_by, 'item_code')
    if direction == 'desc':
        sort_field = f'-{sort_field}'
    
    items = items.order_by(sort_field)

    # 4. Pagination (1000 items per page)
    paginator = Paginator(items, 1000)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'items': page_obj,
        'brand': brand,
        'total_count': paginator.count,
        'current_sort': sort_by,
        'current_direction': direction,
    }

    # AJAX Handling: Return only the rows and pagination info
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('tracking/includes/local_purchase_rows.html', context, request=request)
        pagination_html = render_to_string('tracking/includes/pagination_controls.html', context, request=request)
        return JsonResponse({
            'html': html, 
            'pagination': pagination_html,
            'total_count': paginator.count
        })

    return render(request, 'tracking/local_purchase_list.html', context)

@login_required
@admin_required
def local_purchase_upload(request):
    """View to upload multi-sheet Excel file."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import os
        from django.core.files.storage import FileSystemStorage
        
        excel_file = request.FILES['excel_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)
        
        try:
            with pd.ExcelFile(file_path) as xls:
                sheet_names = xls.sheet_names
                
                # ALLOWED SHEETS WhiteList
                ALLOWED_SHEETS = ['HEPWORTH', 'RAKTherm', 'VERA-PUMP', 'PEGLER', 'OTHERS']
                
                total_imported = 0
                
                for sheet_name in sheet_names:
                    # Filter strictly by allowed names (case-insensitive check)
                    # We check if the sheet name (upper) exists in our allowed list (upper)
                    if sheet_name not in ALLOWED_SHEETS and sheet_name.upper() not in [s.upper() for s in ALLOWED_SHEETS]:
                        continue
                    
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    
                    # Normalize headers: replace newlines and multiple spaces with single space
                    import re
                    df.columns = [re.sub(r'\s+', ' ', str(col)).strip() for col in df.columns]
                    
                    # Check column existence
                    if 'CODE' not in df.columns:
                        continue 
                        
                    # Delete existing for this brand (using the sheet name as brand)
                    LocalPurchaseItem.objects.filter(brand=sheet_name).delete()
                    
                    items_to_create = []
                    for _, row in df.iterrows():
                        def get_val(col, default=0, type_func=int):
                            val = row.get(col)
                            if pd.isna(val) or val == '':
                                return default
                            try:
                                # Handle string inputs like " - " or "N/A"
                                if isinstance(val, str) and not val.replace('.', '', 1).isdigit():
                                    return default
                                return type_func(val)
                            except:
                                return default
                        
                        item = LocalPurchaseItem(
                            brand=sheet_name,
                            item_code=str(row.get('CODE', '')).strip(),
                            upc_code=str(row.get('UPC CODE', '')) if not pd.isna(row.get('UPC CODE')) else '',
                            description=str(row.get('DESCRIPTION', '')) if not pd.isna(row.get('DESCRIPTION')) else '',
                            
                            current_stock_ras=get_val('Current Stock RAS'),
                            current_stock_dip=get_val('Current Stock DIP'),
                            sold_qty_2024=get_val('Sold Qty 2024'),
                            
                            contg=get_val('CONTG.'),
                            trdg=get_val('TRDG.'),
                            stores=get_val('STORES'),
                            
                            total_sold_qty_2025=get_val('TOTAL Sold Qty 2025'),
                            avg_15day_sales=get_val('Avg 15Day Sales HO 2025', 0.0, float),
                            stock_sufficiency_months=get_val('STOCK Sufficiency Month', 0.0, float),
                            lpo_given=get_val('LPO Given'),
                            open_so_qty=get_val('OPEN SO QTY'),
                            stock_reqt_calcn=get_val('STOCK Reqt Calcn'),
                            stock_requirement=get_val('STOCK REQUIREMENT'),
                            value=get_val('VALUE', 0.0, float),
                            cost=get_val('COST', 0.0, float),
                            ho_per_lpo_qty=get_val('HO PER LPO QTY', 0.0, float),
                            stock_reqt_ras_stores=get_val('Stock Reqt RAS/ Stores', 0.0, float),
                        )
                        items_to_create.append(item)
                        
                    LocalPurchaseItem.objects.bulk_create(items_to_create)
                    total_imported += len(items_to_create)
            
            messages.success(request, f"Successfully imported {total_imported} items from {len(sheet_names)} sheets.")
            
        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)
                
        return redirect('local_purchase_dashboard')
        
    return render(request, 'tracking/local_purchase_upload.html')

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect based on role
                try:
                    if not hasattr(user, 'profile'):
                        # Auto-create if missing (self-healing)
                        from .models import UserProfile
                        UserProfile.objects.create(user=user, role='SALESMAN') # Default to Salesman or safe default
                        # Reload user
                        user.refresh_from_db()

                    if user.is_superuser:
                         return redirect('dashboard')
                         
                    if hasattr(user, 'profile') and user.profile.role == 'SALESMAN':
                        return redirect('sales_dashboard')
                    else:
                        # Admin or no profile defaults to dashboard
                        return redirect('dashboard')
                except Exception:
                    # Fallback
                    return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'tracking/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')
